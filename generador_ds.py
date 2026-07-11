# empieza codigo
import streamlit as st
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime
import pandas as pd
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import textwrap
from PIL import Image as PILImage|


################## CTG DISCONNECTOR SWITCH// SECCIONADOR
def mostrar_app():

    st.title("📄 Generador de Ficha CTG")
    st.subheader("Seccionador")
    # 1. Fabricante
    fabricante = "Indicar"
    st.markdown(f" Fabricante: **{fabricante}**")
    # 2. País
    pais = "Indicar"
    st.markdown(f" País: **{pais}**")
    # 3. Referencia
    referencia = "Indicar"
    st.markdown(f" Referencia: **{referencia}**")
    # 4. Norma de fabricación
    norma_fabricacion = "IEC 62271-102"
    st.markdown(f"Norma de fabricación: **{norma_fabricacion}**")
    # 5. Norma de calidad
    norma_calidad = "ISO 9001"
    st.markdown(f"Norma de calidad: **{norma_calidad}**")
    #6. Número de polos
    num_polos = "3"
    st.markdown(f"Número de polos: **{num_polos}**")

    # 7. Instalación
    instalacion = st.selectbox("Tipo de ejecución", ["Exterior", "Interior"])

    # 8. Tipo de accionamiento
    accionamiento = st.selectbox("Tipo de accionamiento", ["Monopolar", "Tripolar"])

    # 9. Tipo de construcción para seccionador de conexión
    conexion = st.selectbox("Tipo de construcción para seccionador de conexión", ["Pantógrafo", "Semi-pantógrafo", "Rotación Central"])
    
    # 10. Altura 
    altura_instalacion = st.number_input("Altura de instalación (m.s.n.m)", min_value=0, value=1000)

    # 11. Temperatura de operación
    st.markdown("### 🌡️ Temperatura de operación")
    temp_min = -5
    st.markdown(f"Temperatura mínima anual (°C): **{temp_min}**")
    temp_max = +40
    st.markdown(f"Temperatura máxima anual (°C): **{temp_max}**")
    temp_media = +35
    st.markdown(f"Temperatura media (24 h) (°C): **{temp_media}**")

    # 12. Frecuencia
    frecuencia_asignada = "60"
    st.markdown(f"Frecuencia asignada (fr) [Hz]: **{frecuencia_asignada}**")

    #13. Clafisicación ambiente sitio de instalación para corrosión según ISO 12944
    corrosion ="Indicar"
    st.markdown(f"Clafisicación ambiente sitio de instalación para corrosión según ISO 12944: **{corrosion}**")

    #14. Nivel de polución sitio de instalación según IEC 60815
    polucion = "Indicar"
    st.markdown(f"Nivel de polución sitio de instalación según IEC 60815: **{polucion}**")

    # 15. Tensión asignada Ur
    ur = st.selectbox("Tensión asignada (Ur) [kV]", options=["145", "245", "550"])

     # 16. Tensión asignada a frecuencia industrial
    # Asignación automática de Ud según Ur
    ud_por_ur = {
        "145": {"fase_tierra_ud": "275", "distancia_seccionamiento": "315"},
        "245": {"fase_tierra_ud": "460", "distancia_seccionamiento": "530"},
        "550": {"fase_tierra_ud": "620", "distancia_seccionamiento": "800"}
    }
    ud_valores = ud_por_ur.get(ur,{"fase_tierra_ud": "", "distancia_seccionamiento": ""})
    st.markdown("### Tensión asignada soportada a frecuencia industrial (Ud)")
    st.markdown(f"a) A tierra y entre polos: **{ud_valores['fase_tierra_ud']}**")
    st.markdown(f"b) A través de la distancia de seccionamiento: **{ud_valores['distancia_seccionamiento']}**")

    # 17. Tensión asignada a impulso maniobra
    # Asignación automática de Us por componente según Ur
    us_por_ur = {
        "145": {"fase_tierra_us": "N.A.", "entre_polos_us": "N.A.", "distancia_seccionamiento_us": "N.A."},
        "245": {"fase_tierra_us": "N.A.", "entre_polos_us": "N.A.", "distancia_seccionamiento_us": "N.A."},
        "550": {"fase_tierra_us": "1175", "entre_polos_us": "1760", "distancia_seccionamiento_us": "900(+450)"}
    }
    us_valores = us_por_ur.get(ur, {"fase_tierra_us": "", "entre_polos_us": "", "distancia_seccionamiento_us": ""})
    st.markdown("#### Tensión asignada soportada a impulso de maniobra (Us)")
    st.markdown(f"a) A tierra: **{us_valores['fase_tierra_us']}**")
    st.markdown(f"b) Entre polos: **{us_valores['entre_polos_us']}**")
    st.markdown(f"c) A través de la distancia de seccionamiento: **{us_valores['distancia_seccionamiento_us']}**")

     # 18. Tensión asignada a impulso tipo rayo
    # Asignación automática de Up según Ur
    up_por_ur = {
        "145": {"fase_tierra_up": "650", "distancia_seccionamiento_up": "750"},
        "245": {"fase_tierra_up": "1050", "distancia_seccionamiento_up": "1200"},
        "550 kV": {"fase_tierra_up": "1500", "distancia_seccionamiento_up": "1500(+315)"}
    }
    up_valores = up_por_ur.get(ur, {"fase_tierra_up": "", "distancia_seccionamiento_up": ""})
    st.markdown("#### Tensión asignada soportada a impulso tipo rayo (Up)")
    st.markdown(f"a) A tierra y entre polos: **{up_valores['fase_tierra_up']}**")
    st.markdown(f"b) A través de la distancia de seccionamiento: **{up_valores['distancia_seccionamiento_up']}**")

    # 19. Corriente asignada (Ir) - Mostrar como texto fijo
    ir_por_ur = {
        "145": "1250",
        "245 kV": "2500",
        "550 kV": "2500"
    }
    # Obtener el valor según la tensión asignada (Ur)
    ir = ir_por_ur.get(ur, "Indicar")
    # Mostrar como texto fijo en la interfaz
    st.markdown(f"Corriente asignada en servicio continuo (Ir): **{ir}**")


    # 20. Corriente de corta duración admisible asignada (Ics)
    ics_por_ur = {
        "145": ["25", "31.5", "40"],
        "245": ["40"],
        "550": ["50"]
    }
    opciones_ics = ics_por_ur.get(ur, [])
    
    if ur == "145 kV":
        ics = st.selectbox("Poder de corte asignado en cortocircuito (Ics) [kA]", opciones_ics)
    else:
        # Mostrar como texto fijo para 245 kV y 550 kV
        ics = opciones_ics[0] if opciones_ics else "Indicar"
        st.markdown(f"###### ⚡ Poder de corte asignado en cortocircuito (Ics): **{ics}**")

    
    # 20. Duración del cortocircuito asignado (Ics)
    duracion_ics = "1"
    st.markdown(f"Duración del cortocircuito asignado (Ics) [s]: **{duracion_ics}**")


    # 21. Corriente de soportabilidad pico asignada (lp) 
    corriente_lp_ur = {
        "145": "82",
        "245": "104",
        "550": "104"
    }
    # Obtener el valor según la tensión asignada (Ur)
    corriente_lp = corriente_lp_ur.get(ur, "Indicar")
    st.markdown(f" Corriente de soportabilidad pico asignada (lp) [kA]: **{corriente_lp}**")

    # 22. Corrientes de transferencia de barras
    st.markdown("### 🔁 Corrientes de transferencia de barras")
    
    # Definición de valores según Ur
    transferencia_por_ur = {
        "145": {"corriente_transferencia": "≥ 0.8 In", "tension_transferencia": "100"},
        "245": {"corriente_transferencia": "≥ 0.6 In", "tension_transferencia": "220"},
        "550": {"corriente_transferencia": "≥ 0.6 In", "tension_transferencia": "220"}
    }
    # Obtener valores según Ur seleccionada
    transferencia_valores = transferencia_por_ur.get(ur, {"corriente_transferencia": "", "tension_transferencia": ""})
    # Mostrar en pantalla
    st.markdown(f"a) Corriente de transferencia en barras asignada [A]: **{transferencia_valores['corriente_transferencia']}**")
    st.markdown(f"b) Tensión de transferencia en barras asignada [V]: **{transferencia_valores['tension_transferencia']}**")

    # 23. Desempeño mecánico mínimo
    desempeno_mecanico = "M2"
    st.markdown(f"### Desempeño mecánico mínimo: **{desempeno_mecanico}**")

    # 24. Distancia mínima en aire
    distancia_entre_polos = "Indicar"
    distancia_a_tierra = "Indicar"
    distancia_seccionamiento = "Indicar"
    
    st.markdown("### Distancia mínima en aire")
    st.markdown(f"a) Entre polos: **{distancia_entre_polos}**")
    st.markdown(f"b) A tierra: **{distancia_a_tierra}**")
    st.markdown(f"c) A través de la distancia de seccionamiento: **{distancia_seccionamiento}**")


    # 25. Aisladores de soporte

    # a) Tipo IEC 60273
    tipo_iec = "Indicar"
    st.markdown(f"Tipo de aislador según IEC 60273: **{tipo_iec}**")
    
    # b) Clase de severidad de contaminación del sitio (SPS)
    st.markdown("### Clase de severidad de contaminación del sitio (SPS) según IEC 60815")
    sps_opciones = {
        "Bajo": 16,
        "Medio": 20,
        "Pesado": 25,
        "Muy Pesado": 31
    }
    sps_seleccion = st.selectbox("Selecciona la clase SPS", list(sps_opciones.keys()))
    valor_sps = sps_opciones[sps_seleccion]
    
    # c) Distancia mínima de fuga requerida
    st.markdown("### 📏 Distancia mínima de fuga requerida")
    um_valores = {"145": 145, "245": 245, "550": 550}
    um_num = um_valores.get(ur, 0)
    distancia_fuga = um_num * valor_sps
    st.text(f"Distancia mínima de fuga: {distancia_fuga} mm")

    # 26. Capacitancia
    capacitancia_contactos_abiertos = "Indicar"
    capacitancia_contactos_tierra = "Indicar"
    
    st.markdown("### ⚡ Capacitancia")
    st.markdown(f"a) Entre contactos abiertos: **{capacitancia_contactos_abiertos}**")
    st.markdown(f"b) Entre contactos y tierra: **{capacitancia_contactos_tierra}**")
    
    # 27. Datos sísmicos
    desempeno_sismico = st.selectbox("Desempeño sísmico según IEEE-693-Vigente", ["Moderado", "Alto"])
    
    # Valores fijos para frecuencia y amortiguamiento
    frecuencia_vibracion = "Indicar"
    coef_amortiguamiento = "Indicar"

    st.markdown("### 🌍 Datos sísmicos")
    st.markdown(f"- Desempeño sísmico según IEEE-693-Vigente: **{desempeno_sismico}**")
    st.markdown(f"a) Frecuencia natural de vibración: **{frecuencia_vibracion}**")
    st.markdown(f"b) Coeficiente de amortiguamiento crítico: **{coef_amortiguamiento}**")


    # 28. Valor asignado de esfuerzo máximo exigido por la maniobra manual
    esfuerzo_seccionador_conexion = "Indicar"
    esfuerzo_seccionador_tierra = "Indicar"
    
    st.markdown("### 🛠️ Valor asignado de esfuerzo máximo exigido por la maniobra manual")
    st.markdown(f"a) Seccionador de conexión: **{esfuerzo_seccionador_conexion}**")
    st.markdown(f"b) Seccionador de puesta a tierra: **{esfuerzo_seccionador_tierra}**")

    # 29. Cargas admisibles en bornes
    cargas_por_ur = {
        "145": {"estatica": "1000", "dinamica": "3000"},
        "245": {"estatica": "1500", "dinamica": "4000"},
        "550": {"estatica": "2000", "dinamica": "5500"}
    }
    cargas_valores = cargas_por_ur.get(ur, {"estatica": "Indicar", "dinamica": "Indicar"})
    carga_estatica = cargas_valores["estatica"]
    carga_dinamica = cargas_valores["dinamica"]

    st.markdown("### 🧱 Cargas admisibles en bornes")
    st.markdown(f"a) Carga estática admisible [N]: **{carga_estatica}**")
    st.markdown(f"b) Carga dinámica admisible [N]: **{carga_dinamica}**")


    # 30. Cuchilla de puesta a tierra
    st.markdown("### ⚙️ Cuchilla de puesta a tierra")
    cuchilla_tierra = st.selectbox("¿Incluye cuchilla de puesta a tierra?", ["Sí", "No"])
    
    if cuchilla_tierra == "Sí":
        # a) Suicheo de corrientes inducidas
        st.markdown("#### a) Suicheo de corrientes inducidas en seccionadores de puesta a tierra")
        clase_suicheo = st.selectbox("Clase de suicheo", ["B", "N.A."])
        corriente_inductiva = "Indicar"
        corriente_capacitiva = "Indicar"
        st.markdown(f"- Corriente inductiva asignada: **{corriente_inductiva}**")
        st.markdown(f"- Corriente capacitiva asignada: **{corriente_capacitiva}**")
    
        # b) Desempeño eléctrico
        st.markdown("#### b) Desempeño eléctrico de seccionadores de puesta a tierra")
        desempeno_electrico = st.selectbox("Desempeño eléctrico", ["E0", "N.A."])
    
        # c) Desempeño mecánico mínimo
        st.markdown("#### c) Desempeño mecánico mínimo")
        desempeno_mecanico_tierra = st.selectbox("Desempeño mecánico mínimo", ["M1", "N.A."])
    else:
        clase_suicheo = "N.A."
        corriente_inductiva = "N.A."
        corriente_capacitiva = "N.A."
        desempeno_electrico = "N.A."
        desempeno_mecanico_tierra = "N.A."

    # 31. Accesorios
    st.markdown("### 🧩 Accesorios")
    # a) Mecanismo de bloqueo externo
    bloqueo_externo = st.selectbox(
        "¿Incluye mecanismo de bloqueo externo y condena en abierto/cerrado?",
        ["Sí", "N.A."]
    )
    # b) Juego de contracontactos
    contracontactos = st.selectbox(
        "¿Incluye juego de contracontactos?",
        ["Sí", "N.A."]
    )

    # 32. Espesor del recubrimiento de plata en contactos principales
    espesor_plata = "Indicar"
    st.markdown(f"### 🧪 Espesor del recubrimiento de plata en contactos principales: **{espesor_plata}**")
    
    # BOTÓN PARA GENERAR FICHA
    ficha_cb = {
        "Fabricante": fabricante,
        "País": pais,
        "Referencia": referencia,
        "Norma de fabricación": norma_fabricacion,
        "Norma de calidad": norma_calidad,
        "Número de polos": num_polos,
        "Instalación": instalacion,
        "Tipo de construcción para seccionador de conexión": conexion,
        "Tipo de accionamiento": accionamiento,
        "Altura de instalación (m.s.n.m)": altura_instalacion,
        "Temperatura mínima anual (°C)": temp_min,
        "Temperatura máxima anual (°C)": temp_max,
        "Temperatura media (24 h) (°C)": temp_media,
        "Frecuencia asignada": frecuencia_asignada,
        "Clafisicación ambiente sitio de instalación para corrosión según ISO 12944": corrosion,
        "Nivel de polución sitio de instalación según IEC 60815": polucion,
        "Tensión asignada (Ur)": ur,
        "Ud - A tierra y entre polos [kV]": ud_valores["fase_tierra_ud"],
        "Ud - A través de la distancia de seccionamiento [kV]": ud_valores["distancia_seccionamiento"],
        "Us - A tierra [kV]": us_valores["fase_tierra_us"],
        "Us - Entre polos [kV]": us_valores["entre_polos_us"],
        "Us - A través de la distancia de seccionamiento [kV]": us_valores["distancia_seccionamiento_us"],
        "Up - A tierra y entre polos [kV]": up_valores["fase_tierra_up"],
        "Up - A través de la distancia de seccionamiento [kV]": up_valores["distancia_seccionamiento_up"],
        "Corriente asignada en servicio continuo (Ir)": ir,
        "Corriente de corta duración admisible asignada (Ics)": ics,
        "Duración del cortocircuito asignado (Ics)": duracion_ics,
        "Corriente de soportabilidad pico asignada (lp)": corriente_lp,
        "Corriente de transferencia en barras asignada": transferencia_valores["corriente_transferencia"],
        "Tensión de transferencia en barras asignada": transferencia_valores["tension_transferencia"],
        "Desempeño mecánico mínimo": desempeno_mecanico,
        "Distancia mínima en aire - Entre polos": distancia_entre_polos,
        "Distancia mínima en aire - A tierra": distancia_a_tierra,
        "Distancia mínima en aire - A través de la distancia de seccionamiento": distancia_seccionamiento,
        "Aislador - Tipo IEC 60273": tipo_iec,
        "Aislador - Clase SPS": sps_seleccion,
        "Aislador - Valor SPS (mm/kV)": valor_sps,
        "Aislador - Distancia mínima de fuga (mm)": distancia_fuga,
        "Capacitancia - Entre contactos abiertos": capacitancia_contactos_abiertos,
        "Capacitancia - Entre contactos y tierra": capacitancia_contactos_tierra,
        "Desempeño sísmico según IEEE-693-Vigente": desempeno_sismico,
        "Frecuencia natural de vibración": frecuencia_vibracion,
        "Coeficiente de amortiguamiento crítico": coef_amortiguamiento,
        "Esfuerzo máximo - Seccionador de conexión": esfuerzo_seccionador_conexion,
        "Esfuerzo máximo - Seccionador de puesta a tierra": esfuerzo_seccionador_tierra,
        "Carga estática admisible en bornes": carga_estatica,
        "Carga dinámica admisible en bornes": carga_dinamica,
        "Cuchilla de puesta a tierra": cuchilla_tierra,
        "Clase de suicheo de corrientes inducidas": clase_suicheo,
        "Corriente inductiva asignada": corriente_inductiva,
        "Corriente capacitiva asignada": corriente_capacitiva,
        "Desempeño eléctrico seccionadores de puesta a tierra": desempeno_electrico,
        "Desempeño mecánico mínimo (puesta a tierra)": desempeno_mecanico_tierra,
        "Accesorio - Mecanismo de bloqueo externo": bloqueo_externo,
        "Accesorio - Juego de contracontactos": contracontactos,
        "Espesor del recubrimiento de plata en contactos principales": espesor_plata
        
    }



    # 📤 Función para exportar Excel con estilo personalizado
    def exportar_excel(datos, fuente="Calibri", tamaño=9):
        # Diccionario de unidades (puedes ampliarlo según tus campos)
        unidades = {
            "Tensión asignada (Ur) [kV]": "kV",
            "Altura de instalación (m.s.n.m)": "m.s.n.m",
            "Temperatura mínima anual (°C)": "°C",
            "Temperatura máxima anual (°C)": "°C",
            "Temperatura media (24 h) (°C)": "°C",
            "Frecuencia asignada": "Hz",
            "Corriente asignada en servicio continuo (Ir)": "A",
            "Poder de corte asignado en cortocircuito (Ics)": "kA",
            "Duración del cortocircuito asignado (Ics)": "s",
            "Porcentaje de corriente aperiódica (%)": "%",
            "Distancia mínima en aire - Entre polos (mm)": "mm",
            "Distancia mínima de fuga (mm)": "mm",
            "Campo eléctrico a 1 metro de separación del piso (kV/m)": "kV/m",
            "Masa neta para transporte (kg)": "kg",
            "Volumen total para transporte (m³)": "m³",
            "Dimensiones para transporte (Alto x Ancho x Largo) [mm]": "mm",
            "Masa neta de un polo completo con estructura (kg)": "kg",

            "Ud - A tierra y entre polos [kV]": "kV",
            "Ud - A través de la distancia de seccionamiento [kV]": "kV",
            "Us - A tierra [kV]": "kV",
            "Us - Entre polos [kV]": "kV",
            "Us - A través de la distancia de seccionamiento [kV]": "kV",
            "Up - A tierra y entre polos [kV]": "kV",
            "Up - A través de la distancia de seccionamiento [kV]": "kV",

            "Corriente de soportabilidad pico asignada (lp)": "kA",
            "Corriente de transferencia en barras asignada": "A",
            "Tensión de transferencia en barras asignada": "V",
            "Desempeño mecánico mínimo": "Clase",
            "Distancia mínima en aire - Entre polos": "mm",
            "Distancia mínima en aire - A tierra": "mm",
            "Distancia mínima en aire - A través de la distancia de seccionamiento": "mm",
            "Aislador - Distancia mínima de fuga (mm)": "mm",
            "Capacitancia - Entre contactos abiertos": "pF",
            "Capacitancia - Entre contactos y tierra": "pF",
 
            "Frecuencia natural de vibración": "Hz",
            "Coeficiente de amortiguamiento crítico": "%",
            "Esfuerzo máximo - Seccionador de conexión": "N",
            "Esfuerzo máximo - Seccionador de puesta a tierra": "N",
            "Carga estática admisible en bornes": "N",
            "Carga dinámica admisible en bornes": "N",
            "Corriente inductiva asignada": "A",
            "Corriente capacitiva asignada": "A",
            "Desempeño eléctrico seccionadores de puesta a tierra": "Clase",
            "Desempeño mecánico mínimo (puesta a tierra)": "Clase",
            "Espesor del recubrimiento de plata en contactos principales": "μm"
            
        }

      # Crear DataFrame con estructura personalizada
        df = pd.DataFrame([
            {
                "ÍTEM": i + 1,
                "DESCRIPCIÓN": campo,
                "UNIDAD": unidades.get(campo, ""),
                "REQUERIDO": valor,
                "OFRECIDO": ""  # Columna vacía para completar manualmente
            }
            for i, (campo, valor) in enumerate(datos.items())
        ])
    
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="CTG", startrow=6)
            wb = writer.book
            ws = writer.sheets["CTG"]
            ws.print_title_rows = '1:7'
            ws.print_area = f"A1:E{ws.max_row}"

            
            # 🖼️ Insertar imagen del logo (opcional)
            logo_path = "siemens_logo.png"
            try:
                # Abrimos con PIL para procesar cualquier formato (incluyendo webp disfrazado)
                img_aux = PILImage.open(logo_path)
                img_converted = BytesIO()
                img_aux.save(img_converted, format="PNG") # Forzamos PNG
                img_converted.seek(0)
                
                # Insertamos en Excel
                img = Image(img_converted)
                img.width = 280
                img.height = 90
                ws.add_image(img, "C1")
            except Exception as e:
                st.warning(f"⚠️ No se pudo insertar el logo: {e}. El archivo se generará sin imagen.")
    
            # 🟪 Caja de título
            ws.merge_cells("A2:E4")
            cell = ws.cell(row=2, column=1)
            cell.value = "FICHA TÉCNICA SECCIONADOR"
            cell.font = Font(name=fuente, bold=True, size=14, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
            # 🏷️ Subtítulo técnico
            ws.merge_cells("A5:D5")
            ws["A5"] = f"CARACTERÍSTICAS GARANTIZADAS"
            ws["A5"].font = Font(name=fuente, bold=True, size=12)
            ws["A5"].alignment = Alignment(horizontal="center")
    
            # 🎨 Encabezados con estilo
            header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            header_font = Font(name=fuente, size=tamaño, color="FFFFFF", bold=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
            for col_num in range(1, 6):
                cell = ws.cell(row=6, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
    
            # 📐 Ajuste de columnas
            ws.column_dimensions["A"].width = 4
            ws.column_dimensions["B"].width = 50
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 12
    
            
            
            # 📋 Formato de filas con fuente personalizada y ajuste dinámico de altura
            for row in ws.iter_rows(min_row=7, max_row=ws.max_row, max_col=5):
                max_lines = 1  # Mínimo una línea por celda
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.font = Font(name=fuente, size=tamaño)
            
                    # Estimar número de líneas necesarias si el contenido es texto
                    if cell.value and isinstance(cell.value, str):
                        # Ajusta el ancho según la columna (por ejemplo, columna B tiene 55 caracteres de ancho)
                        if cell.column_letter == "B":
                            wrapped = textwrap.wrap(cell.value, width=55)
                            max_lines = max(max_lines, len(wrapped))
            
                # Ajustar altura de la fila según el contenido más largo
                ws.row_dimensions[row[0].row].height = max_lines * 15  # 15 puntos por línea aprox.
            
                # Alineación horizontal para columnas específicas
                row[0].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                row[2].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                row[3].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                row[4].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
                
        output.seek(0)
        return output
    
    # 📥 Botón para generar y descargar
    fuente = "Calibri"
    tamaño = 9
    if st.button("📊 Generar archivo CTG"):
        archivo_excel = exportar_excel(ficha_cb, fuente=fuente, tamaño=tamaño)
        nivel_tension = ficha_cb.get("Nivel de tensión (kV)", "XX")
        st.download_button(
            label="📥 Descargar archivo CTG en Excel",
            data=archivo_excel,
            file_name=f"CTG_{nivel_tension}kV.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
            




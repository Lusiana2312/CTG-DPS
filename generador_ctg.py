
#empieza codigo
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

st.set_page_config(page_title="Generador CTG", layout="wide")
st.title("📄 Generador de Archivos CTG")

# 🛠️ Parámetros editables
with st.expander("🛠️ Parámetros editables"):
    nivel_tension = st.selectbox("Nivel de tensión (kV):", [115, 230, 500])
    # Diccionario de Ur según nivel de tensión
    ur_por_nivel = {
        115: 100,
        230: 200,
        500: 400
    }

    # Obtener Ur automáticamente
    ur = ur_por_nivel[nivel_tension]
    st.markdown(f"🔌 Tensión asignada (Ur): **{ur} kV**")

    altura_instalacion = st.number_input("Altura sobre el nivel del mar (m.s.n.m):", min_value=0, value=1000)

    sps_opciones = {
        "Bajo": 16,
        "Medio": 20,
        "Pesado": 25,
        "Muy Pesado": 31
    }
    sps_seleccion = st.selectbox("Clase de severidad de contaminación del sitio (SPS):", list(sps_opciones.keys()))
    valor_sps = sps_opciones[sps_seleccion]
    st.markdown(f"🔢 Valor SPS seleccionado: **{valor_sps}**")

    ka = st.number_input("Coeficiente Ka:", min_value=1.0, max_value=2.0, value=1.0, step=0.1)
    km = st.number_input("Coeficiente Km:", min_value=1.0, max_value=2.0, value=1.0, step=0.1)

    desempeno_sismico = st.selectbox("Desempeño sísmico vigente:", ["Alto", "Moderado"])

    distancia_fuga = nivel_tension * valor_sps * ka * km
    st.markdown(f"📏 **Distancia mínima de fuga requerida:** {distancia_fuga:.2f} mm")

# 📘 Parámetros definidos por norma
with st.expander("📘 Parámetros definidos por norma"):
    datos_definidos = {
        "Norma de fabricación": "IEC 60099-4",
        "Norma de calidad": "IEC 9001",
        "Tipo de ejecución": "Exterior",
        "Frecuencia asignada (f)": "60",
        "Material cubierta": "Polimérico",
        "Número de columnas": "1",
        "Tensión más elevada para el material (Um)": "245",
        "Tensión continua de operación (Uc)": "20",
        "Corriente de descarga asignada (In)": "40",
        "Clase de descarga de línea": "4",
        "Capacidad mínima de disipación de energía asignada para dos impulsos de larga duración kJ/kV (Ur)": "≥10",
        "Transferencia de carga repetitiva Qrs": "≥2.4",
        "Contador de descargas": "Sí"
    }
    for campo, valor in datos_definidos.items():
        st.markdown(f"**{campo}:** {valor}")

# ⚡ Tensiones residuales
with st.expander("⚡ Tensiones residuales"):
    tensiones_residuales = {
        "Tensión residual al impulso de corriente de escalón (10 kA)": "",
        "Tensión residual al impulso tipo maniobra (250 A)": "",
        "Tensión residual al impulso tipo maniobra (500 A)": "",
        "Tensión residual al impulso tipo maniobra (1000 A)": "",
        "Tensión residual al impulso tipo maniobra (2000 A)": "",
        "Tensión residual al impulso tipo rayo (5 kA)": "",
        "Tensión residual al impulso tipo rayo (10 kA)": "",
        "Tensión residual al impulso tipo rayo (20 kA)": "",
        "Tensión asignada soportada a la frecuencia industrial (Ud)": "",
        "Tensión asignada soportada al impulso tipo rayo (Up)": "",
        "Tensión asignada soportada al impulso tipo maniobra (Us)": ""
    }
    for campo in tensiones_residuales:
        tensiones_residuales[campo] = st.text_input(campo, value="")

# 📋 Consolidar todos los datos
datos = {
    "Nivel de tensión (kV)": nivel_tension,
    "Tensión asignada (Ur)": ur,
    "Altura de instalación (m.s.n.m)": altura_instalacion,
    "Clase SPS": sps_seleccion,
    "Valor SPS": valor_sps,
    "Coeficiente Ka": ka,
    "Coeficiente Km": km,
    "Distancia mínima de fuga (mm)": round(distancia_fuga, 2),
    "Desempeño sísmico vigente": desempeno_sismico
}
datos.update(datos_definidos)
datos.update(tensiones_residuales)



# 📤 Función para exportar Excel con estilo personalizado
def exportar_excel(datos, fuente="Calibri", tamaño=9):
    unidades = {
        "Nivel de tensión (kV)": "kV",
        "Tensión asignada (Ur)": "kV",
        "Altura de instalación (m.s.n.m)": "m.s.n.m",
        "Coeficiente Ka": "",
        "Coeficiente Km": "",
        "Distancia mínima de fuga (mm)": "mm",
        "Tensión residual al impulso de corriente de escalón (10 kA)": "kV",
        "Tensión residual al impulso tipo maniobra (250 A)": "kV",
        "Tensión residual al impulso tipo maniobra (500 A)": "kV",
        "Tensión residual al impulso tipo maniobra (1000 A)": "kV",
        "Tensión residual al impulso tipo maniobra (2000 A)": "kV",
        "Tensión residual al impulso tipo rayo (5 kA)": "kV",
        "Tensión residual al impulso tipo rayo (10 kA)": "kV",
        "Tensión residual al impulso tipo rayo (20 kA)": "kV",
        "Tensión asignada soportada a la frecuencia industrial (Ud)": "kV",
        "Tensión asignada soportada al impulso tipo rayo (Up)": "kV",
        "Tensión asignada soportada al impulso tipo maniobra (Us)": "kV"
    }

    df = pd.DataFrame([
        {
            "ÍTEM": i + 1,
            "DESCRIPCIÓN": campo,
            "UNIDAD": unidades.get(campo, ""),
            "REQUERIDO": valor,
            "OFRECIDO": "" #Columna vacía
        }
        for i, (campo, valor) in enumerate(datos.items())
    ])

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="CTG", startrow=6)
        wb = writer.book
        ws = writer.sheets["CTG"]

        # 🖼️ Insertar imagen del logo
        logo_path = "siemens_logo.png"
        try:
            img = Image(logo_path)
            img.width = 300
            img.height = 100
            ws.add_image(img, "C1")
        except FileNotFoundError:
            st.warning("⚠️ No se encontró el logo 'siemens_logo.png'. Asegúrate de subirlo al repositorio.")
        
        #🧱 Crear borde negro alrededor de A2:E4
        black_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        for row in ws.iter_rows(min_row=2, max_row=4, min_col=1, max_col=5):
            for cell in row:
                cell.border = black_border

        # 🟪 Caja de título
        ws.merge_cells("A2:E4")
        cell = ws.cell(row=2, column=1)
        cell.value = "CARACTERÍSTICAS GARANTIZADAS"
        cell.font = Font(name=fuente, bold=True, size=14, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # 🏷️ Título técnico
        ws.merge_cells("A5:D5")
        ws["A5"] = f"DESCARGADORES DE SOBRETENSIÓN {datos['Nivel de tensión (kV)']} kV"
        ws["A5"].font = Font(name=fuente, bold=True, size=12)
        ws["A5"].alignment = Alignment(horizontal="center")

        # 🎨 Encabezados con estilo
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_font = Font(name=fuente, size=tamaño, color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_num in range(1, 6):
            cell = ws.cell(row=6, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # 📐 Ajuste de columnas
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12

        # 📋 Formato de filas con fuente personalizada
        for row in ws.iter_rows(min_row=7, max_row=ws.max_row, max_col=5):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.font = Font(name=fuente, size=tamaño)
            row[0].alignment = Alignment(horizontal="center", vertical="center")
            row[2].alignment = Alignment(horizontal="center", vertical="center")
            row[3].alignment = Alignment(horizontal="center", vertical="center")
            row[4].alignment = Alignment(horizontal="center", vertical="center")  # Alineación para OFRECIDO


    output.seek(0)
    return output


# 📥 Botón para generar y descargar
fuente = "Calibri"
tamaño = 9
if st.button("📊 Generar archivo CTG"):
    archivo_excel = exportar_excel(datos, fuente=fuente, tamaño=tamaño)
    nivel_tension = datos.get("Nivel de tensión (kV)", "XX")
    st.download_button(
        label="📥 Descargar archivo CTG en Excel",
        data=archivo_excel,
        file_name=f"CTG_{nivel_tension}kV.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )













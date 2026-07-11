# empieza codigo
import streamlit as st
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime
import numpy as np
import pandas as pd
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import textwrap
import math
from PIL import Image as PILImage

def mostrar_app():

    st.title("📄 Generador de Ficha CTG")
    st.subheader("Transformador de Tensión")

    st.markdown("### ⚙️ Parámetros del transformador")

    # 1 al 5
    fabricante = pais = referencia = "Indicar"
    st.text("1. Fabricante: " + fabricante)
    st.text("2. País: " +  pais)
    st.text("3. Referencia: " + referencia)
    norma_fabricacion = "IEC 61869-5"
    st.text("4. Norma de fabricación: " + norma_fabricacion)
    norma_calidad = "ISO 9001"
    st.text("5. Norma de calidad: " + norma_calidad)

    # 6 al 9
    tipo_ejecucion = st.selectbox("6. Tipo de ejecución", ["Interior", "Exterior"])
    altura_instalacion = st.number_input("7. Altura de instalación (msnm)", min_value=0, step=100)
    material_aislador = st.selectbox("8. Material del aislador", ["Compuesto siliconado", "Porcelana"])
    tipo_transformador = st.selectbox("8a. Tipo", ["Capacitivo", "Inductivo"])
    tension_um = st.selectbox("9. Tensión más elevada para el material (Um)", ["123 kV", "245 kV", "550 kV"])

    # 10. Ud
    st.markdown("### 🔌 10. Tensión asignada soportada a la frecuencia industrial (Ud)")
    ud_interno = {"145 kV": "360 kV", "245 kV": "460 kV", "550 kV": "700 kV"}[tension_um]
    st.text(f"Tensión asignada soportada a la frecuencia industrial (Ud)-Aislamiento Interno a condiciones normales de prueba: {ud_interno}")
    st.text(f"Tensión asignada soportada a la frecuencia industrial (Ud)-Aislamiento Externo a condiciones normales de prueba (*): {ud_interno} a {int(altura_instalacion)} msnm")

    # 11. Up
    st.markdown("### ⚡ 11. Tensión asignada soportada al impulso tipo rayo (Up)")
    up_interno = {"145 kV": "750 kV", "245 kV": "1050 kV", "550 kV": "1550 kV"}[tension_um]
    st.text(f"Tensión asignada soportada al impulso tipo rayo (Up)-Aislamiento Interno a condiciones normales de prueba: {up_interno}")
    st.text(f"Tensión asignada soportada al impulso tipo rayo (Up)-Aislamiento Externo a condiciones normales de prueba (*): {up_interno} a {int(altura_instalacion)} msnm")

    # 12. Us
    st.markdown("### ⚡ 12. Tensión asignada soportada al impulso tipo maniobra (Us)")
    us_interno = st.text_input("Tensión asignada soportada al impulso tipo maniobra (Us)-Aislamiento Interno a condiciones normales de prueba (dejar vacío por ahora)")
    us_externo = st.text_input("Tensión asignada soportada al impulso tipo maniobra (Us)-Aislamiento Externo a condiciones normales de prueba (*) (dejar vacío por ahora)")

    # 13. Frecuencia
    st.markdown("### 📶 13. Frecuencia asignada (fr)")
    st.text("60 Hz")

    # 14. factor de tensión asignado
    st.markdown("### ⚙️ 14. Factor de tensión asignado")
    # Opción a) Permanente
    factor_permanente = "1,2"
    st.text("a) Permanente: " + factor_permanente)
    # Opción b) Durante 30 s
    factor_30s = "1,5"
    st.text("b) Durante 30 s: " + factor_30s)

    # 15. Capacidad total
    if tension_um == "145 kV":
        capacidad_minima = 2000
    elif tension_um == "245 kV":
        capacidad_minima = 4000
    elif tension_um == "550 kV":
        capacidad_minima = 10000
    else:
        capacidad_minima = 0  # Fallback por si se añade otra opción
    # Campo de entrada para capacidad total
    st.markdown("### ⚡ 15. Capacidad total")
    capacidad_total = st.number_input(
        f"Capacidad total (≥ {capacidad_minima} pF)",
        min_value=capacidad_minima
    )
    
    # 16 al 18
    st.markdown("### 🔧 16-18. Condensadores y tensión intermedia")
    c1 = c2 = tension_intermedia = "Indicar"
    st.text("16. Condensador de alta tensión (C1): " + c1)
    st.text("17. Condensador de tensión intermedia (C2): " + c2)
    st.text("18. Tensión intermedia asignada en circuito abierto: " + tension_intermedia)

    # 19. Número de devanados secundarios
    st.markdown("### 🔁 19. Número de devanados secundarios")
    num_devanados = st.selectbox("Selecciona el número de devanados secundarios", [1, 2, 3])

    # 20. Clase de precisión
    st.markdown("### 🎯 20. Clase de precisión")
    st.markdown("**Entre el 25% y el 100% de la carga de precisión con factor de potencia 0,8 en atraso**")
    clase_precision_a = clase_precision_c = "05-3P"
    st.text("a) Entre el 5% y el 80% de la tensión asignada: " + clase_precision_a)
    clase_precision_b = "0,2"
    st.text("b) Entre el 80% y el 120% de la tensión asignada: " + clase_precision_b)
    st.text("c) Entre el 120% y el 150% de la tensión asignada: " + clase_precision_c)

    # 21. Carga de precisión
    st.markdown("### ⚙️ 21. Carga de precisión")
    rango_burden = st.selectbox("Rango de burden acorde con IEC 61869-1/3/5", ["I", "II", "III", "IV"])
    st.text("a) Devanado 1: 15 VA")
    st.text("b) Devanado 2: 15 VA")
    st.text("c) Devanado 3: 15 VA")
    st.text("d) Simultánea: 45 VA")
    potencia_termica_limite = "Indicar"
    st.text("e) Potencia térmica límite: " + potencia_termica_limite)


    # 22. Tensión asignada  
    st.markdown("### ⚡ 22. Tensión asignada")
    # a) Tensión primaria
    upn_opciones = [110, 230, 500]
    upn_seleccionada = st.selectbox("a) Tensión primaria (Upn)", upn_opciones)
    # Mostrar solo la expresión simbólica
    st.text(f"{upn_seleccionada} / √3")
    
    # b) Tensión secundaria
    usn_opciones = ["115 / √3", "110 / √3"]
    usn_seleccionada = st.selectbox("b) Tensión secundaria (Usn)", usn_opciones)
    
    # Mostrar la expresión seleccionada
    st.text(usn_seleccionada)

    # 23. Distancia de arco
    st.markdown("### 🧯 23. Distancia de arco")
    distancia_arco = "Indicar"
    st.text("Distancia de arco (mm): " + distancia_arco)

    # Distancia mínima de fuga
    st.markdown("### 📏 26. Distancia mínima de fuga requerida")
    
    # Selección de clase SPS
    sps_opciones = {"Bajo": 16, "Medio": 20, "Pesado": 25, "Muy Pesado": 31}
    sps_seleccion = st.selectbox("Selecciona la clase SPS", list(sps_opciones.keys()))
    valor_sps = sps_opciones[sps_seleccion]
    
    # Conversión de tensión Um a valor numérico
    um_valores = {"145 kV": 145, "245 kV": 245, "550 kV": 550}
    um_num = um_valores.get(tension_um, 0)

    distancia_fuga = um_num * valor_sps
    st.text(f"Distancia mínima de fuga: {distancia_fuga} mm")

    # 24. Dispositivos de Protección: (Todos los modelos contarán con accesorios carrier)
    disp_prot = st.selectbox(f"Dispositivos de Protección: (Todos los modelos contarán con accesorios carrier): ", ["Sí", "No"])

    # Cuchilla externa
    cuchilla_ext = st.selectbox(f"a)Cuchilla externa para cortocircuitar la parte inductiva del equipo del lado del divisor capacitivo (Ground Potencial Switch): ", ["Sí", "No"])

    # Dispositivo amortiguador
    disp_amort = st.selectbox(f"b) Dispositivo amortiguador de ferrorresonancia: ", ["Sí", "No"])

    # Interruptores miniatura
    int_mini = st.selectbox(f"c) Interruptores miniatura adecuados para protección de los circuitos secundarios de tensión, con contactos auxiliares para indicación de apertura y disparo. La curva de operación de los interruptores miniatura debe ser del Tipo Z: ", ["Sí", "No"])

    # 25. Accesorios
    #placa de características
    placa = st.selectbox(f"a) Placa de características de acuerdo con lo estipulado en las Publicaciones IEC 61869-3 e IEC 61869-5. Dentro de la placa se deberá indicar que el aceite es libre de PCB y azufre corrosivo.", ["Sí", "No"])
    
    # BOTÓN PARA GENERAR FICHA
    # 📋 Diccionario con los datos del transformador
    ficha_ctg = {
        "Fabricante": fabricante,
        "País": pais,
        "Referencia": referencia,
        "Norma de fabricación": norma_fabricacion,
        "Norma de calidad": norma_calidad,
        "Tipo de ejecución": tipo_ejecucion,
        "Altura de instalación (msnm)": altura_instalacion,
        "Material del aislador": material_aislador,
        "Tipo de transformador": tipo_transformador,
        "Tensión más elevada para el material (Um)": tension_um,
        "Tensión asignada soportada a la frecuencia industrial (Ud)-Aislamiento Interno a condiciones normales de prueba": ud_interno,
        "Tensión asignada soportada a la frecuencia industrial (Ud)-Aislamiento Externo a condiciones normales de prueba (*)": f"{ud_interno} a {int(altura_instalacion)} msnm",
        "Tensión asignada soportada al impulso tipo rayo (Up)-Aislamiento Interno a condiciones normales de prueba": up_interno,
        "Tensión asignada soportada al impulso tipo rayo (Up)-Aislamiento Externo a condiciones normales de prueba (*)": f"{up_interno} a {int(altura_instalacion)} msnm",
        "Tensión asignada soportada al impulso tipo maniobra (Us)-Aislamiento Interno a condiciones normales de prueba": us_interno,
        "Tensión asignada soportada al impulso tipo maniobra (Us)-Aislamiento Externo a condiciones normales de prueba (*)": us_externo,
        "Frecuencia asignada (fr)": "60 Hz",
        "Factor de tensión permanente": str(factor_permanente),
        "Factor de tensión durante 30 s": str(factor_30s),
        "Capacidad total (VA)": capacidad_total,
        "Condensador de alta tensión (C1)": c1,
        "Condensador de tensión intermedia (C2)": c2,
        "Tensión intermedia en circuito abierto": tension_intermedia,
        "Número de devanados secundarios": num_devanados,
        "Clase de precisión (5%-80%)": clase_precision_a,
        "Clase de precisión (80%-120%)": clase_precision_b,
        "Clase de precisión (120%-150%)": clase_precision_c,
        "Rango de burden (IEC 61869)": rango_burden,
        "Carga Devanado 1 (VA)": "15",
        "Carga Devanado 2 (VA)": "15",
        "Carga Devanado 3 (VA)": "15",
        "Carga Simultánea (VA)": "45",
        "Potencia térmica límite": potencia_termica_limite,
        "Tensión primaria (Upn)": f"{upn_seleccionada} / √3",
        "Tensión secundaria (Usn)": usn_seleccionada,
        "Distancia de arco (mm)": distancia_arco,
        "Distancia mínima de fuga (mm)": distancia_fuga,
        "Dispositivos de Protección: (Todos los modelos contarán con accesorios carrier)": disp_prot,
        "a)Cuchilla externa para cortocircuitar la parte inductiva del equipo del lado del divisor capacitivo (Ground Potencial Switch)": cuchilla_ext,
        "b) Dispositivo amortiguador de ferrorresonancia": disp_amort,
        "c) Interruptores miniatura adecuados para protección de los circuitos secundarios de tensión, con contactos auxiliares para indicación de apertura y disparo. La curva de operación de los interruptores miniatura debe ser del Tipo Z":int_mini,
        "Accesorios":"",
        "a) Placa de características de acuerdo con lo estipulado en las Publicaciones IEC 61869-3 e IEC 61869-5. Dentro de la placa se deberá indicar que el aceite es libre de PCB y azufre corrosivo.":placa    
    }

    #Función exportar Excel con estilo personalizado
    def exportar_excel(datos, fuente="Calibri", tamaño=9):
        unidades = {
            "Altura de instalación (msnm)": "msnm",
            "Capacidad total (VA)": "VA",
            "Tensión más elevada para el material (Um)": "kV",
            "Tensión Ud - Aislamiento Interno": "kV",
            "Tensión Ud - Aislamiento Externo": "kV",
            "Tensión Up - Aislamiento Interno": "kV",
            "Tensión Up - Aislamiento Externo": "kV",
            "Tensión Us - Aislamiento Interno": "kV",
            "Tensión Us - Aislamiento Externo": "kV",
            "Frecuencia asignada (fr)": "Hz",
            "Factor de tensión permanente": "",
            "Factor de tensión durante 30 s": "",
            "Carga Devanado 1 (VA)": "VA",
            "Carga Devanado 2 (VA)": "VA",
            "Carga Devanado 3 (VA)": "VA",
            "Carga Simultánea (VA)": "VA",
            "Tensión primaria (Upn)": "V",
            "Tensión secundaria (Usn)": "V"
            # Puedes añadir más unidades si lo deseas
        }
    
        df = pd.DataFrame([
            {
                "ÍTEM": i + 1,
                "DESCRIPCIÓN": campo,
                "UNIDAD": unidades.get(campo, ""),
                "REQUERIDO": valor,
                "OFRECIDO": ""
            }
            for i, (campo, valor) in enumerate(datos.items())
        ])
    
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="CTG", startrow=6)
            wb = writer.book
            ws = writer.sheets["CTG"]
    
            # 🖼️ Logo (opcional)
            logo_path = "siemens_logo.png"
            try:
                # Abrimos con PIL para procesar cualquier formato (incluyendo webp disfrazado)
                img_aux = PILImage.open(logo_path)
                img_converted = BytesIO()
                img_aux.save(img_converted, format="PNG") # Forzamos PNG
                img_converted.seek(0)
                
                # Insertamos en Excel
                img = Image(img_converted)
                img.width = 280
                img.height = 90
                ws.add_image(img, "C1")
            except Exception as e:
                st.warning(f"⚠️ No se pudo insertar el logo: {e}. El archivo se generará sin imagen.")

            
            # 🟪 Título
            ws.merge_cells("A2:E4")
            cell = ws.cell(row=2, column=1)
            cell.value = "FICHA TÉCNICA TRANSFORMADOR DE TENSIÓN"
            cell.font = Font(name=fuente, bold=True, size=14, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
            # 🏷️ Subtítulo
            ws.merge_cells("A5:D5")
            ws["A5"] = "CARACTERÍSTICAS GARANTIZADAS"
            ws["A5"].font = Font(name=fuente, bold=True, size=12)
            ws["A5"].alignment = Alignment(horizontal="center")
    
            # 🎨 Encabezados
            header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            header_font = Font(name=fuente, size=tamaño, color="FFFFFF", bold=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
            for col_num in range(1, 6):
                cell = ws.cell(row=6, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
            # 📐 Ajuste de columnas
            ws.column_dimensions["A"].width = 4
            ws.column_dimensions["B"].width = 50
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 12
    
            # 📋 Formato de filas
            for row in ws.iter_rows(min_row=7, max_row=ws.max_row, max_col=5):
                max_lines = 1
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.font = Font(name=fuente, size=tamaño)
    
                    if cell.value and isinstance(cell.value, str):
                        if cell.column_letter == "B":
                            wrapped = textwrap.wrap(cell.value, width=55)
                            max_lines = max(max_lines, len(wrapped))
    
                ws.row_dimensions[row[0].row].height = max_lines * 15
                row[0].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                row[2].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                row[3].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                row[4].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
        output.seek(0)
        return output

    fuente="Calibri"
    tamaño=9
    if st.button("📊 Generar archivo CTG"):
        archivo_excel = exportar_excel(ficha_ctg, fuente=fuente, tamaño=tamaño)
        st.download_button(
            label="📥 Descargar archivo CTG en Excel",
            data=archivo_excel,
            file_name="CTG_Transformador_Tension.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )







































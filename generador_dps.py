
#empieza codigo
import streamlit as st
import pandas as pd 		
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
    
def mostrar_app():
    st.title("📄 Generador de Archivos CTG")

    # 1. Altura de instalación
    altura_instalacion = st.number_input("### 🧱 Altura sobre el nivel del mar (m.s.n.m)", min_value=0, value=1000, step=10)

    # 2. Fabricante (valor fijo)
    fabricante = "Indicar"
    st.text("### 🏢 Fabricante: " + fabricante)

    # 3. Referencia (valor fijo)
    referencia = "Indicar"
    st.text("### 🏷️ Referencia: " + referencia)

    # 4. Norma de fabricación (fijo)
    norma_fabricacion = "IEC 60099-4"
    st.text("### 📘 Norma de fabricación: " + norma_fabricacion)

    # 5. Norma de calidad
    norma_calidad = st.selectbox("### 📘 Norma de calidad", ["IEC 9001", "ISO 9001"])

    # 6. Tipo de ejecución
    tipo_ejecucion = st.selectbox("### 🏗️ Tipo de ejecución", ["Exterior", "Interior"])

    # 7. Frecuencia asignada (fijo)
    frecuencia_asignada = "60 Hz"
    st.text("### ⚡ Frecuencia asignada: " + frecuencia_asignada)

    # 8. Material cubierta
    material_cubierta = st.selectbox("### 🧩 Material de la cubierta", ["Polimérico", "Porcelana"])

    # 9. Número de columnas (valor fijo)
    numero_columnas = 1
    st.text(f"### 🔢 Número de columnas: {numero_columnas}")

    # 10. Número de cuerpos
    numero_cuerpos = st.text_input("### 🔢 Número de cuerpos", value="Indicar")

    # 11. Tensión más elevada para el material (Um)
    um = st.selectbox("### ⚡ Tensión más elevada para el material (Um)", ["145 kV", "245 kV", "550 kV"])

    # 12. Tensión asignada (Ur)
    ur_por_um = {"145 kV": "110 kV", "245 kV": "198 kV", "550 kV": "444 kV"}
    ur = ur_por_um.get(um, "")
    st.text(f"### ⚡ Tensión asignada (Ur): {ur}")

    # 13. Tensión continua de operación (Uc)
    uc ="Indicar"
    st.text("### ⚡ Tensión continua de operación (Uc): " + uc)

    # 14. Corriente de descarga asignada (In) - valor fijo
    in_corriente = "20 kA"
    st.text(f"### ⚡ Corriente de descarga asignada (In): {in_corriente}")

    # 15. Corriente asignada del dispositivo de alivio de presión (0.2 seg) - valor fijo
    corriente_alivio = "40 kA"
    st.text(f"### ⚡ Corriente asignada del dispositivo de alivio de presión (0.2 seg): {corriente_alivio}")

    # 16. Tensión residual al impulso de corriente de escalón (10 kA)
    ures_escalon= "Indicar"
    st.text("### ⚡ Tensión residual al impulso de corriente de escalón (10 kA): " +ures_escalon)

    # 17. Tensión residual al impulso tipo maniobra (Ures)
    st.markdown("### ⚡ Tensión residual al impulso tipo maniobra (Ures)")
    ures_maniobra_250 = "Indicar"
    st.text("Ures - Para 250 A: " + ures_maniobra_250)
    ures_maniobra_500 = "Indicar"
    st.text("Ures - Para 500 A: " + ures_maniobra_500)
    ures_maniobra_1000 = "Indicar"
    st.text("Ures - Para 1000 A: " + ures_maniobra_1000)
    ures_maniobra_2000 = "Indicar"
    st.text("Ures - Para 2000 A: " + ures_maniobra_2000 )

    # 18. Tensión residual al impulso tipo rayo (Ures)
    st.markdown("### ⚡ Tensión residual al impulso tipo rayo (Ures)")
    ures_rayo_5ka = "Indicar"
    st.text("Ures - 5 kA: " + ures_rayo_5ka)
    ures_rayo_10ka = "Indicar"
    st.text("Ures - 10 kA: " + ures_rayo_10ka)
    ures_rayo_20ka = "Indicar"
    st.text("Ures - 20 kA: " + ures_rayo_20ka)

    # 19. Clase de descarga de línea (automática según Um)
    clase_por_um = {"145 kV": 3, "245 kV": 4, "550 kV": 5}
    clase_descarga = clase_por_um.get(um, "No definida")
    st.text(f"### ⚡ Clase de descarga de línea: {clase_descarga}")

    # 20. Capacidad mínima de disipación de energía
    capacidad_duracion = "≥10 kJ/kV"
    st.text(f"### ⚡ Capacidad mínima de disipación de energía (2 impulsos largos): {capacidad_duracion}")

    # 21. Transferencia de carga repetitiva Qrs
    qrs = "≥2.4"
    st.text(f"### ⚡ Transferencia de carga repetitiva Qrs: {qrs}")

    # 22. Mínima sobretensión temporal soportada
    st.markdown("### ⚡ Mínima sobretensión temporal soportada luego de absorber la energía asignada")
    sobretension_1s = "Indicar"
    st.text("Durante 1s: " + sobretension_1s)
    sobretension_10s = "Indicar"
    st.text("Durante 10s: " + sobretension_10s)

    # 23. Capacitancia fase-tierra
    capacitancia = "Indicar"
    st.text("### ⚡ Capacitancia fase-tierra: " + capacitancia)

    # 24. Distancia de arco
    distancia_arco = "Indicar"
    st.text("### ⚡ Distancia de arco (con anillos anticorona si aplica): " + distancia_arco)

    # 25. Clase de severidad de contaminación del sitio (SPS)
    st.markdown("### 🌫️ Clase de severidad de contaminación del sitio (SPS)")
    sps_opciones = {"Bajo": 16, "Medio": 20, "Pesado": 25, "Muy Pesado": 31}
    sps_seleccion = st.selectbox("Selecciona la clase SPS", list(sps_opciones.keys()))
    valor_sps = sps_opciones[sps_seleccion]

    # 26. Distancia mínima de fuga = Um * SPS
    st.markdown("### 📏 Distancia mínima de fuga requerida")
    um_valores = {"145 kV": 123, "245 kV": 245, "550 kV": 550}
    um_num = um_valores.get(um, 0)
    distancia_fuga = um_num * valor_sps
    st.text(f"Distancia mínima de fuga: {distancia_fuga} mm")

    # 27. Aislamiento de la envolvente
    st.markdown("### 🧪 Aislamiento de la envolvente (con anillos anticorona si aplica)")
    ud = "Indicar"
    st.text("Tensión asignada soportada a la frecuencia industrial (Ud): " + ud)
    up = "Indicar"
    st.text("Tensión asignada soportada al impulso tipo rayo (Up): " + up)
    us = "Indicar"
    st.text("Tensión asignada soportada al impulso tipo maniobra (Us): " + us)

    # 28. Datos sísmicos
    st.markdown("### 🌍 Datos sísmicos según IEEE-693 vigente")
    desempeno_sismico = st.selectbox("Desempeño sísmico", ["Bajo", "Moderado (0.25g)", "Alto (0.5g)"])
    frecuencia_natural = "Indicar"
    st.text("Frecuencia natural de vibración: " + frecuencia_natural)
    amortiguamiento_critico = "Indicar"
    st.text("Coeficiente de amortiguamiento crítico: " + amortiguamiento_critico)

    # 29. Cargas admisibles en bornes (automáticas según Um)
    st.markdown("### 🧱 Cargas admisibles en bornes")
    cargas_por_um = {
        "145 kV": {"estatica": "500 N", "dinamica": "1000 N"},
        "245 kV": {"estatica": "1000 N", "dinamica": "2000 N"},
        "550 kV": {"estatica": "2000 N", "dinamica": "5000 N"}
    }
    cargas = cargas_por_um.get(um, {"estatica": "No definida", "dinamica": "No definida"})
    
    # ✅ Definir las variables antes de mostrarlas
    carga_estatica = cargas["estatica"]
    carga_dinamica = cargas["dinamica"]
    
    st.text(f"Carga estática admisible: {carga_estatica}")
    st.text(f"Carga dinámica admisible: {carga_dinamica}")

    # 30. Altura total
    altura_total = "Indicar"
    st.text("### 📏 Altura total: " + altura_total)

    # 31. Dimensiones para transporte
    dimensiones_transporte = "Indicar"
    st.text("### 📦 Dimensiones para transporte (Alto x Ancho x Largo): " + dimensiones_transporte)

    # 32. Masa neta para transporte
    masa_transporte = "Indicar"
    st.text("### ⚖️ Masa neta para transporte: " + masa_transporte)

    # 33. Volumen total
    volumen_total = "Indicar"
    st.text("### 📦 Volumen total" + volumen_total)

    # 34. Anillo corona y de distribución de campo
    anillo_corona = "Indicar"
    st.text("### 🧲 Anillo corona y de distribución de campo: " + anillo_corona)

    st.markdown("### 🧰 Accesorios")
    # 35. Contador de descargas
    contador_descargas = st.selectbox("### 🔌 Contador de descargas", ["Sí", "No"])



    
    
    # Consolidar todos los datos en un solo diccionario
    datos = {
        # 1–3
        "Altura sobre el nivel del mar (m.s.n.m)": altura_instalacion,
        "Fabricante": fabricante,
        "Referencia": referencia,
    
        # 4–7
        "Norma de fabricación": norma_fabricacion,
        "Norma de calidad": norma_calidad,
        "Tipo de ejecución": tipo_ejecucion,
        "Frecuencia asignada (Hz)": frecuencia_asignada,
    
        # 8–10
        "Material de la cubierta": material_cubierta,
        "Número de columnas": numero_columnas,
        "Número de cuerpos": numero_cuerpos,
    
        # 11–12
        "Tensión más elevada para el material (Um)": um,
        "Tensión asignada (Ur)": ur,
    
        # 13–15
        "Tensión continua de operación (Uc)": uc,
        "Corriente de descarga asignada (In)": in_corriente,
        "Corriente asignada del dispositivo de alivio de presión (0.2 seg)": corriente_alivio,
    
        # 16
        "Tensión residual al impulso de corriente de escalón (10 kA)": ures_escalon,

        #17
        
        "250 A\n": ures_maniobra_250,
        "Tensión residual al impulso tipo maniobra (Ures) - 500 A\n": ures_maniobra_500,
        "Tensión residual al impulso tipo maniobra (Ures) - 1000 A\n": ures_maniobra_1000,
        "Tensión residual al impulso tipo maniobra (Ures) - 2000 A": ures_maniobra_2000,

        # 18
        "Tensión residual al impulso tipo rayo (Ures) - 5 kA": ures_rayo_5ka,
        "Tensión residual al impulso tipo rayo (Ures) - 10 kA": ures_rayo_10ka,
        "Tensión residual al impulso tipo rayo (Ures) - 20 kA": ures_rayo_20ka,
    
        # 19–21
        "Clase de descarga de línea": clase_descarga,
        "Capacidad mínima de disipación de energía (kJ/kV)": capacidad_duracion,
        "Transferencia de carga repetitiva Qrs": qrs,
    
        # 22
        "Sobretensión temporal soportada - 1s": sobretension_1s,
        "Sobretensión temporal soportada - 10s": sobretension_10s,
    
        # 23–24
        "Capacitancia fase-tierra": capacitancia,
        "Distancia de arco (con anillos anticorona)": distancia_arco,
    
        # 25–26
        "Clase SPS": sps_seleccion,
        "Valor SPS": valor_sps,
        "Distancia mínima de fuga (mm)": distancia_fuga,
    
        # 27
        "Tensión soportada a frecuencia industrial (Ud)": ud,
        "Tensión soportada al impulso tipo rayo (Up)": up,
        "Tensión soportada al impulso tipo maniobra (Us)": us,
    
        # 28
        "Desempeño sísmico": desempeno_sismico,
        "Frecuencia natural de vibración": frecuencia_natural,
        "Coeficiente de amortiguamiento crítico": amortiguamiento_critico,
    
        # 29
        "Carga estática admisible": carga_estatica,
        "Carga dinámica admisible": carga_dinamica,
    
        # 30–34
        "Altura total": altura_total,
        "Dimensiones para transporte (Alto x Ancho x Largo)": dimensiones_transporte,
        "Masa neta para transporte": masa_transporte,
        "Volumen total": volumen_total,
        "Anillo corona y de distribución de campo": anillo_corona,
    
        # 35–36
        "Accesorios": "",
        "Contador de descargas": contador_descargas
    }
        
        
    # 📤 Función para exportar Excel con estilo personalizado
    def exportar_excel(datos, fuente="Calibri", tamaño=9):
        unidades = {
            "Altura sobre el nivel del mar (m.s.n.m)": "m.s.n.m",
            "Tensión asignada (Ur)": "kV",
            "Tensión más elevada para el material (Um)": "kV",
            "Tensión continua de operación (Uc)": "kV",
            "Corriente de descarga asignada (In)": "kA",
            "Corriente asignada del dispositivo de alivio de presión (0.2 seg)": "kA",
            "Tensión residual al impulso de corriente de escalón (10 kA)": "kV",
            "Tensión residual al impulso tipo maniobra (Ures) - 250 A": "kV",
            "Tensión residual al impulso tipo maniobra (Ures) - 500 A": "kV",
            "Tensión residual al impulso tipo maniobra (Ures) - 1000 A": "kV",
            "Tensión residual al impulso tipo maniobra (Ures) - 2000 A": "kV",
            "Tensión residual al impulso tipo rayo (Ures) - 5 kA": "kV",
            "Tensión residual al impulso tipo rayo (Ures) - 10 kA": "kV",
            "Tensión residual al impulso tipo rayo (Ures) - 20 kA": "kV",
            "Distancia mínima de fuga (mm)": "mm",
            "Tensión soportada a frecuencia industrial (Ud)": "kV",
            "Tensión soportada al impulso tipo rayo (Up)": "kV",
            "Tensión soportada al impulso tipo maniobra (Us)": "kV"
        }
    
        df = pd.DataFrame([
            {
                "ÍTEM": i + 1,
                "DESCRIPCIÓN": campo,
                "UNIDAD": unidades.get(campo, ""),
                "REQUERIDO": valor,
                "OFRECIDO": ""
            }
            for i, (campo, valor) in enumerate(datos.items())
        ])

        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="CTG", startrow=6)
            wb = writer.book
            ws = writer.sheets["CTG"]
            # Repetir las filas de título (A1 a E5) en cada página impresa
            ws.print_title_rows = '1:7'
            
            # Definir el área de impresión completa desde A1 hasta la última fila de la tabla
            ws.print_area = f"A1:E{ws.max_row}"
    
            # Logo
            logo_path = "siemens_logo.png"
            try:
                # Abrimos con PIL para limpiar cualquier formato raro (.webp)
                img_aux = PILImage.open(logo_path)
                img_converted = BytesIO()
                img_aux.save(img_converted, format="PNG") # Forzamos formato PNG real
                img_converted.seek(0)
            
                # Ahora openpyxl recibe un PNG garantizado
                img_for_excel = Image(img_converted)
                img_for_excel.width = 280
                img_for_excel.height = 90
                ws.add_image(img_for_excel, "C1")
            except Exception as e:
                st.warning(f"⚠️ El logo no pudo ser insertado (Formato incompatible). El Excel se generará sin logo.")

            # Borde negro
            black_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            for row in ws.iter_rows(min_row=2, max_row=4, min_col=1, max_col=5):
                for cell in row:
                    cell.border = black_border
    
            # Título principal
            ws.merge_cells("A2:E4")
            cell = ws.cell(row=2, column=1)
            cell.value = "CARACTERÍSTICAS GARANTIZADAS"
            cell.font = Font(name=fuente, bold=True, size=14, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
            # Título técnico
            tension_texto = datos.get("Tensión asignada (Ur)", "XX")
            ws.merge_cells("A5:D5")
            ws["A5"] = f"DESCARGADORES DE SOBRETENSIÓN {tension_texto}"
            ws["A5"].font = Font(name=fuente, bold=True, size=12)
            ws["A5"].alignment = Alignment(horizontal="center")
    
            # Encabezados
            header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            header_font = Font(name=fuente, size=tamaño, color="FFFFFF", bold=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            for col_num in range(1, 6):
                cell = ws.cell(row=6, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
    
            # Ajuste de columnas
            ws.column_dimensions["A"].width = 4
            ws.column_dimensions["B"].width = 50
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 12
    
            # Formato de filas
            for row in ws.iter_rows(min_row=7, max_row=ws.max_row, max_col=5):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.font = Font(name=fuente, size=tamaño)
                row[0].alignment = Alignment(horizontal="center", vertical="center")
                row[2].alignment = Alignment(horizontal="center", vertical="center")
                row[3].alignment = Alignment(horizontal="center", vertical="center")
                row[4].alignment = Alignment(horizontal="center", vertical="center")
    
        output.seek(0)
        return output
        
        
    # 📥 Botón para generar y descargar
    if st.button("📊 Generar archivo CTG"):
        archivo_excel = exportar_excel(datos, fuente="Calibri", tamaño=9)
        tension_texto = datos.get("Tensión asignada (Ur)", "XX")
        st.download_button(
            label="📥 Descargar archivo CTG en Excel",
            data=archivo_excel,
            file_name=f"CTG_{tension_texto}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        
        
        
        
        
        
    
    
    
    
    














































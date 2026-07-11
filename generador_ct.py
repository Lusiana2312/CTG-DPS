# empieza codigo
import streamlit as st
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime
import pandas as pd
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import textwrap
from PIL import Image as PILImage


def mostrar_app():


    st.title("📄 Generador de Ficha CTG")
    st.subheader("Transformador de Corriente")

    # 1. DATOS GENERALES
    st.markdown("### 🖊️ Datos generales")
    fabricante = st.text_input("Fabricante")
    pais = st.text_input("País")
    referencia = st.text_input("Referencia")
    norma_fabricacion = "IEC 61869-1 y IEC 61869-2"
    norma_calidad = "ISO 9001"
    st.markdown(f"**Norma de fabricación:** {norma_fabricacion}")
    st.markdown(f"**Norma de calidad:** {norma_calidad}")

    tipo_ejecucion = st.selectbox("Tipo de ejecución", options=["Exterior", "Interior"])
    altura_instalacion = st.number_input("Altura de instalación [msnm]", min_value=0)
    material_aislador = st.selectbox("Material del aislador", options=["Compuesto Siliconado", "Material"])

    # 2. PARÁMETROS DE TENSIÓN
    st.markdown("### ⚡ Parámetros de tensión")
    tension_material = st.selectbox("Tensión más elevada para el material (Um)", options=["145 kV", "245 kV", "550 kV"])

    # Asignación automática de tensiones
    if tension_material == "145 kV":
        ud_interno = "360 kV"
        up_interno = "750 kV"
        ipn = "1000 A"
    elif tension_material == "245 kV":
        ud_interno = "460 kV"
        up_interno = "1050 kV"
        ipn = "2500 A"
    elif tension_material == "550 kV":
        ud_interno = "700 kV"
        up_interno = "1550 kV"
        ipn = "3000 A"
    else:
        ud_interno = ""
        up_interno = ""
        ipn = ""

    ud_externo = f"{ud_interno} a {altura_instalacion} msnm" if ud_interno else ""
    up_externo = f"{up_interno} a {altura_instalacion} msnm" if up_interno else ""

    st.write("**Tensión asignada soportada a la frecuencia industrial (Ud) - Aislamiento Interno:**", ud_interno)
    st.write("**Tensión asignada soportada a la frecuencia industrial (Ud) - Aislamiento Externo:**", ud_externo)
    st.write("**Tensión asignada soportada al impulso tipo rayo (Up) - Aislamiento Interno:**", up_interno)
    st.write("**Tensión asignada soportada al impulso tipo rayo (Up) - Aislamiento Externo:**", up_externo)

    us_interno = st.text_input("Tensión asignada soportada al impulso tipo maniobra (Us) - Aislamiento Interno")
    us_externo = st.text_input("Tensión asignada soportada al impulso tipo maniobra (Us) - Aislamiento Externo")

    # 3. PARÁMETROS ELÉCTRICOS
    st.markdown("### 🔌 Parámetros eléctricos")
    frecuencia = "60 Hz"
    st.markdown(f"**Frecuencia asignada (fr):** {frecuencia}")
    st.markdown(f"**Corriente primaria asignada (Ipn):** {ipn}")
    factor_ipn = "1"
    st.markdown(f"**Factor de la corriente primaria continua asignada:** {factor_ipn}")
    isn = "1"
    st.markdown(f"**Corriente secundaria asignada (Isn):** {isn}")
    ith = "40 kA"
    st.markdown(f"**Corriente de cortocircuito térmica asignada (Ith) en 1 segundo:** {ith}")
    idyn = "2.6 × Ith"
    st.markdown(f"**Corriente dinámica asignada (Idyn):** {idyn}")


    #4. 
    st.markdown("### 🧲 Configuración de núcleos")
    
    # Selección de cantidad de núcleos de medida y protección
    num_medida = st.selectbox("Número de núcleos de medida (máx. 2)", options=[1, 2])
    num_proteccion = st.selectbox("Número de núcleos de protección convencional (máx. 4)", options=[3, 4])
    
    nucleos = []
    
    # Características generales para núcleos de medida
    st.markdown("#### ⚙️ Características generales de núcleos de medida")
    
    relacion_asignada = "2500-1250-625/1"
    relacion_exactitud = "2500-1250-625/1"
    clase_exactitud = "0,2 S"
    carga_exactitud = "Indicar"
    
    st.markdown(f"**a) Relación de transformación asignada:** {relacion_asignada}")
    st.markdown(f"**b) Relación para la que debe cumplir la exactitud:** {relacion_exactitud}")
    st.markdown(f"**c) Clase de exactitud:** {clase_exactitud}")
    st.markdown(f"**d) Carga de exactitud núcleos de medida:** {carga_exactitud}")
    
    # Listado de relaciones y valores por núcleo

    # Listado de relaciones y valores por núcleo

    def obtener_relaciones_por_um(tension_material):
        if tension_material == "145 kV":
            return {
                "300/1  (2S3-2S4)": "7,5",
                "600/1  (2S2-2S4)": "15",
                "1200/1  (2S1-2S4)": "30"
            }
        elif tension_material == "245 kV":
            return {
                "625/1   (1S3-1S4)": "2,5",
                "1250/1  (1S2-1S4)": "5",
                "2500/1  (1S1-1S4)": "15",
                "400/1   (1S3-1S4)": "NA",
                "800/1   (1S2-1S4)": "NA",
                "1600/1  (1S1-1S4)": "NA"
            }
        elif tension_material == "550 kV":
            return {
                "500/1 (3S3-3S4)": "2,5",
                "1000/1 (3S2-3S4)": "5",
                "2000/1 (3S1-3S4)": "15",
                "625/1  (3S3-3S4)": "NA",
                "1250/1  (3S2-3S4)": "NA",
                "2500/1  (3S1-3S4)": "NA"
            }
        else:
            return {}
    
    # Obtener las relaciones según el valor de Um seleccionado
    relaciones_valores = obtener_relaciones_por_um(tension_material)
    
    # Mostrar cada relación como parámetro individual en Streamlit
    for i in range(num_medida):
        st.markdown(f"##### Parámetros individuales para núcleo de medida {i+1}")
        for relacion, valor in relaciones_valores.items():
            st.write(f"{relacion}: {valor}")
    
        # Guardar cada relación como clave individual en el diccionario
        relaciones_individuales = {f"Núcleo {i+1} - {relacion}": valor for relacion, valor in relaciones_valores.items()}
    
        nucleos.append({
            "Número": i + 1,
            "Tipo": "Medida",
            "Relación asignada": relacion_asignada,
            "Relación exactitud": relacion_exactitud,
            "Clase exactitud": clase_exactitud,
            "Carga exactitud": carga_exactitud,
            **relaciones_individuales  # Se agregan las relaciones como claves individuales
        })

    #23. CAMBIO DE RELACIÓN EN EL SECUNDARIO
    st.markdown("### 🔄 Cambio de relación en el secundario")
    cambio_relacion_secundario = st.selectbox("¿Existe cambio de relación en el secundario?", options=["Sí", "No"])

    #24. Dispositivo de protección primario
    st.markdown("### Dispositivo de protección primario")
    
    fabricante_proteccion = "Indicar"
    referencia_proteccion = "Indicar"
    
    st.markdown(f"**a) Fabricante:** {fabricante_proteccion}")
    st.markdown(f"**b) Referencia:** {referencia_proteccion}")
        
    # 7. CAPACIDAD
    st.markdown("### Capacidad")
    capacidad = "Indicar"
    st.markdown(f"**Capacidad:** {capacidad}")

    # 8. DISTANCIA DE ARCO
    st.markdown("### Distancia de arco")
    distancia_arco = "Indicar"
    st.markdown(f"**Distancia de arco:** {distancia_arco}")

    # 9. DISTANCIA MÍNIMA DE FUGA
    st.markdown("### Distancia mínima de fuga")
    
    # Selección de clase SPS
    sps_opciones = {"Bajo": 16, "Medio": 20, "Pesado": 25, "Muy Pesado": 31}
    sps_seleccion = st.selectbox("Selecciona la clase SPS", list(sps_opciones.keys()))
    valor_sps = sps_opciones[sps_seleccion]
    
    um_valores = {"145 kV": 145, "245 kV": 245, "550 kV": 550}
    um_num = um_valores.get(tension_material, 0)
    distancia_fuga = um_num * valor_sps
    st.markdown(f"**Distancia mínima de fuga requerida:** {distancia_fuga} mm")

    # 11. DESEMPEÑO SÍSMICO SEGÚN IEEE-693
    st.markdown("### 🌍 Desempeño sísmico según IEEE-693-Vigente (**)")
    desempeno_sismico = st.selectbox(
        "Selecciona el nivel de desempeño sísmico",
        options=["Moderado (0,25 g)", "Alto (0,5 g)"]
    )
    st.markdown(f"**Desempeño sísmico seleccionado:** {desempeno_sismico}")
    
    frecuencia_vibracion = "Indicar"
    coef_amortiguamiento = "Indicar"
    
    st.markdown(f"**a) Frecuencia natural de vibración:** {frecuencia_vibracion}")
    st.markdown(f"**b) Coeficiente de amortiguamiento crítico:** {coef_amortiguamiento}")

    # 13. CARGAS ADMISIBLES EN BORNES
    st.markdown("### Cargas admisibles en bornes")
    
    # Asignación según Um
    if tension_material == "145 kV":
        carga_estatica = 1000
        carga_dinamica = 3000
    elif tension_material == "245 kV":
        carga_estatica = 1500
        carga_dinamica = 4000
    elif tension_material == "550 kV":
        carga_estatica = 2000
        carga_dinamica = 5500
    else:
        carga_estatica = "Indicar"
        carga_dinamica = "Indicar"
    
    st.markdown(f"**a) Carga estática admisible:** {carga_estatica} N")
    st.markdown(f"**b) Carga dinámica admisible:** {carga_dinamica} N")

    # 11. Temperatura de operación
    st.markdown("### 🌡️ Temperatura de operación")
    temp_min = -10
    st.text(f"### Temperatura mínima anual (°C): {temp_min}")
    temp_max = +40
    st.text(f"### Temperatura máxima anual (°C): {temp_max}")
    temp_media = +35
    st.text(f"### Temperatura media (24 h) (°C): {temp_media}")

    # 14. GRADO DE PROTECCIÓN CAJA DE TERMINALES SECUNDARIA
    st.markdown("### 🧰 Grado de protección caja de terminales secundaria")
    
    grado_proteccion = "IP55"
    st.markdown(f"**Grado de protección:** {grado_proteccion}")

    # 15. CLASIFICACIÓN AMBIENTE SITIO DE INSTALACIÓN PARA CORROSIÓN
    st.markdown("### Clasificación del ambiente del sitio de instalación para corrosión según ISO 12944")
    
    clasificacion_corrosion = "Indicar"
    st.markdown(f"**Clasificación según ISO 12944:** {clasificacion_corrosion}")

    # 16. ACCESORIOS
    st.markdown("### 🧩 Parámetro 33: Accesorios")
    
    accesorio_a = st.selectbox(
        "a) Tapón de sello roscado para llenado de aceite instalado en la parte superior",
        options=["Sí", "No"]
    )
    accesorio_b = st.selectbox(
        "b) Indicadores de nivel de aceite fácilmente visible desde el piso con visor resistente a los rayos UV",
        options=["Sí", "No"]
    )
    accesorio_c = st.selectbox(
        "c) Placa de características de acuerdo con la Norma IEC 61869-2 (incluye indicación de aceite libre de PCB y azufre corrosivo)",
        options=["Sí", "No"]
    )
    accesorio_d = st.selectbox(
        "d) Tap capacitivo para pruebas de factor de potencia",
        options=["Sí", "No"]
    )
    accesorio_e = st.selectbox(
        "e) Bornera fija seccionable en núcleos secundarios tipo URTK/S o similar con accesorio de bloqueo (Phoenix Contact S-0308359 o similar)",
        options=["Sí", "No"]
    )

    # 17. DISPOSITIVO DE PROTECCIÓN CONTRA SOBRETENSIONES
    st.markdown("### ⚡ Parámetro 34: Dispositivo de protección contra sobretensiones")
    
    proteccion_sobretension = st.selectbox(
        "¿El equipo incluye dispositivo de protección contra sobretensiones que limite la tensión a 2500 V sin alterar la exactitud del núcleo?",
        options=["Sí", "No"]
    )

    # 18. TIEMPO DE DETECCIÓN DE LA FALLA
    st.markdown("### ⏱️ Parámetro 35: Tiempo de detección de la falla")
    st.markdown("**Resultado:** ≤ 6 segundos")

    # 19. APLICACIÓN DE NORMA IEC 60255-121
    st.markdown("### 📘 Parámetro 36: Aplicación de norma IEC 60255-121 para protecciones de distancia")
    
    norma_iec_60255 = st.selectbox(
        "¿Se aplica la norma IEC 60255-121 para el dimensionamiento de los núcleos de protecciones de distancia?",
        options=["Sí", "No"]
    )

    # 20. APLICACIÓN DE NORMA IEC 60255-151
    st.markdown("### ⚙️ Parámetro 37: Aplicación de norma IEC 60255-151 para protecciones de sobrecorriente")
    
    norma_iec_60255_151 = st.selectbox(
        "¿Se aplica la norma IEC 60255-151 para el dimensionamiento de los núcleos de protecciones de sobrecorriente?",
        options=["Sí", "No"]
    )

    # 21. CANTIDAD DE ACEITE
    st.markdown("### Cantidad de aceite")
    cantidad_aceite = "Indicar"
    st.markdown(f"**Cantidad de aceite:** {cantidad_aceite}")

    # 22. DIMENSIONES PARA TRANSPORTE
    st.markdown("###Dimensiones para transporte (Alto x Ancho x Largo)")
    dimensiones_transporte = "Indicar"
    st.markdown(f"**Dimensiones para transporte:** {dimensiones_transporte}")

    # 23. MASA NETA PARA TRANSPORTE
    st.markdown("### Masa neta para transporte")
    masa_neta_transporte = "Indicar"
    st.markdown(f"**Masa neta para transporte:** {masa_neta_transporte}")
    
    # 23. Volumen total
    st.markdown("### Volumen total")
    volumen_total = "Indicar"
    st.markdown(f"**Volumen total:** {volumen_total}")

    # 23. Vida útil del equipo
    st.markdown("### Volumen total")
    vida_util = "Indicar"
    st.markdown(f"**Vida útil del equipo:** {vida_util}")
    

    # BOTÓN PARA GENERAR FICHA
    ficha_cb = {
        "Fabricante": fabricante,
        "País": pais,
        "Referencia": referencia,
        "Norma de fabricación": norma_fabricacion,
        "Norma de calidad": norma_calidad,
        "Tipo de ejecución": tipo_ejecucion,
        "Altura de instalación [msnm]": altura_instalacion,
        "Material del aislador": material_aislador,
        "Tensión más elevada para el material (Um)": tension_material,
        "Tensión asignada soportada a la frecuencia industrial (Ud) - Aislamiento Interno a condiciones normales de prueba": ud_interno,
        "Tensión asignada soportada a la frecuencia industrial (Ud) - Aislamiento Externo a condiciones normales de prueba (*)": ud_externo,
        "Tensión asignada soportada al impulso tipo rayo (Up) - Aislamiento Interno a condiciones normales de prueba ": up_interno,
        "Tensión asignada soportada al impulso tipo rayo (Up) - Aislamiento Externo a condiciones normales de prueba  (*)": up_externo,
        "Tensión asignada soportada al impulso tipo maniobra (Us) - Aislamiento Interno a condiciones normales de prueba ": us_interno,
        "Tensión asignada soportada al impulso tipo maniobra (Us) - Aislamiento Externo a condiciones normales de prueba  (*)": us_externo,
        "Frecuencia asignada (fr)": frecuencia,
        "Corriente primaria asignada (Ipn)": ipn,
        "Factor de la corriente primaria contínua asignada ": factor_ipn,
        "Corriente secundaria asignada (Isn)": isn,
        "'Corriente de cortocircuito térmica asignada (Ith) en '1 segundo": ith,
        "Corriente dinámica asignada (Idyn)": idyn,
        "Cantidad y clase de núcleos":"",
        "a) Medida": num_medida,
        "b) Protección convencional": num_proteccion,
        "Características núcleos de medida": "",
        "a) Relación de transformación asignada ": relacion_asignada,
        "b) Relación para la que debe cumplir la exactitud": relacion_exactitud,
        "c) Clase de exactitud": clase_exactitud,
        "d) Carga de exactitud núcleoa de medida": carga_exactitud,
        "-Núcleo 1":"",
        "Relación de transformación asignada": relacion_asignada,
        "Relación para exactitud": relacion_exactitud,
        "Clase de exactitud": clase_exactitud,
        "Carga de exactitud": carga_exactitud,
        "Cambio de relación en el secundario": cambio_relacion_secundario,
        "Dispositivo de protección primario - Fabricante": fabricante_proteccion,
        "Dispositivo de protección primario - Referencia": referencia_proteccion,
        "Capacidad": capacidad,
        "Distancia de arco": distancia_arco,
        "Distancia mínima de fuga requerida (mm)": distancia_fuga,
        "Clase de severidad de contaminación del sitio (SPS)": sps_seleccion,
        "Desempeño sísmico según IEEE-693-Vigente": desempeno_sismico,
        "a - Frecuencia natural de vibración": frecuencia_vibracion,
        "b - Coeficiente de amortiguamiento crítico": coef_amortiguamiento,
        "a - Carga estática admisible (N)": carga_estatica,
        "b - Carga dinámica admisible (N)": carga_dinamica,
        "Relación asignada": relacion_asignada,
        "Relación exactitud": relacion_exactitud,
        "Clase exactitud": clase_exactitud,
        "Carga exactitud": carga_exactitud,
        "Temperatura mínima anual (°C)": temp_min,
        "Temperatura máxima anual (°C)": temp_max,
        "Temperatura media (24 h) (°C)": temp_media,
        "Grado de protección caja de terminales secundaria": grado_proteccion,
        "Clasificación ambiente sitio de instalación para corrosión según ISO 12944": clasificacion_corrosion,
        "a - Tapón de sello roscado para llenado de aceite": accesorio_a,
        "b - Indicadores de nivel de aceite con visor UV": accesorio_b,
        "c - Placa de características IEC 61869-2 (sin PCB ni azufre corrosivo)": accesorio_c,
        "d - Tap capacitivo para pruebas de factor de potencia": accesorio_d,
        "e - Bornera fija seccionable tipo URTK/S con bloqueo": accesorio_e,
        "Protección contra sobretensiones (2500 V sin alterar exactitud)": proteccion_sobretension,
        "Tiempo de detección de la falla": "≤ 6 segundos",
        "Aplicación de norma IEC 60255-121 para protecciones de distancia": norma_iec_60255,
        "Aplicación de norma IEC 60255-151 para protecciones de sobrecorriente": norma_iec_60255_151,
        "Cantidad de aceite": cantidad_aceite,
        "Dimensiones para transporte (Alto x Ancho x Largo)": dimensiones_transporte,
        "Masa neta para transporte": masa_neta_transporte,
        "Volumen total": volumen_total,
        "Vida útil del equipo": vida_util
        

    }
    

    # 📤 Función para exportar Excel con estilo personalizado
    def exportar_excel(datos, fuente="Calibri", tamaño=9):
        # Diccionario de unidades (puedes ampliarlo según tus campos)
        unidades = {
            "Tensión asignada (Ur) [kV]": "kV",
            "Altura de instalación (m.s.n.m)": "m.s.n.m",
            "Temperatura mínima anual (°C)": "°C",
            "Temperatura máxima anual (°C)": "°C",
            "Temperatura media (24 h) (°C)": "°C",
            "Frecuencia asignada (fr)": "Hz",
            "Corriente asignada en servicio continuo (Ir)": "A",
            "Poder de corte asignado en cortocircuito (Ics)": "kA",
            "Duración del cortocircuito asignado (Ics)": "s",
            "Porcentaje de corriente aperiódica (%)": "%",
            "Distancia mínima en aire - Entre polos (mm)": "mm",
            "Distancia mínima de fuga (mm)": "mm",
            "Campo eléctrico a 1 metro de separación del piso (kV/m)": "kV/m",
            "Masa neta para transporte (kg)": "kg",
            "Volumen total para transporte (m³)": "m³",
            "Dimensiones para transporte (Alto x Ancho x Largo) [mm]": "mm",
            "Masa neta de un polo completo con estructura (kg)": "kg"
            # Añade más unidades según tus campos
        }

            
        # Crear DataFrame con estructura personalizada
        df = pd.DataFrame([
            {
                "ÍTEM": i + 1,
                "DESCRIPCIÓN": campo,
                "UNIDAD": unidades.get(campo, ""),
                "REQUERIDO": valor,
                "OFRECIDO": ""  # Columna vacía para completar manualmente
            }
            for i, (campo, valor) in enumerate(datos.items())
        ])
    
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="CTG", startrow=6)
            wb = writer.book
            ws = writer.sheets["CTG"]
            ws.print_title_rows = '1:7'
            ws.print_area = f"A1:E{ws.max_row}"

            
            # 🖼️ Insertar imagen del logo (opcional)
            logo_path = "siemens_logo.png"
            try:
                # Abrimos con PIL para procesar cualquier formato (incluyendo webp disfrazado)
                img_aux = PILImage.open(logo_path)
                img_converted = BytesIO()
                img_aux.save(img_converted, format="PNG") # Forzamos PNG
                img_converted.seek(0)
                
                # Insertamos en Excel
                img = Image(img_converted)
                img.width = 280
                img.height = 90
                ws.add_image(img, "C1")
            except Exception as e:
                st.warning(f"⚠️ No se pudo insertar el logo: {e}. El archivo se generará sin imagen.")
    
            # 🟪 Caja de título
            ws.merge_cells("A2:E4")
            cell = ws.cell(row=2, column=1)
            cell.value = "FICHA TÉCNICA INTERRUPTOR DE POTENCIA"
            cell.font = Font(name=fuente, bold=True, size=14, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
            # 🏷️ Subtítulo técnico
            ws.merge_cells("A5:D5")
            ws["A5"] = f"CARACTERÍSTICAS GARANTIZADAS"
            ws["A5"].font = Font(name=fuente, bold=True, size=12)
            ws["A5"].alignment = Alignment(horizontal="center")
    
            # 🎨 Encabezados con estilo
            header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            header_font = Font(name=fuente, size=tamaño, color="FFFFFF", bold=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
            for col_num in range(1, 6):
                cell = ws.cell(row=6, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
    
            # 📐 Ajuste de columnas
            ws.column_dimensions["A"].width = 4
            ws.column_dimensions["B"].width = 50
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 12
    
            
            
            # 📋 Formato de filas con fuente personalizada y ajuste dinámico de altura
            for row in ws.iter_rows(min_row=7, max_row=ws.max_row, max_col=5):
                max_lines = 1  # Mínimo una línea por celda
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.font = Font(name=fuente, size=tamaño)
            
                    # Estimar número de líneas necesarias si el contenido es texto
                    if cell.value and isinstance(cell.value, str):
                        # Ajusta el ancho según la columna (por ejemplo, columna B tiene 55 caracteres de ancho)
                        if cell.column_letter == "B":
                            wrapped = textwrap.wrap(cell.value, width=55)
                            max_lines = max(max_lines, len(wrapped))
            
                # Ajustar altura de la fila según el contenido más largo
                ws.row_dimensions[row[0].row].height = max_lines * 15  # 15 puntos por línea aprox.
            
                # Alineación horizontal para columnas específicas
                row[0].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                row[2].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                row[3].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                row[4].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
                
        output.seek(0)
        return output
    
    # 📥 Botón para generar y descargar
    fuente = "Calibri"
    tamaño = 9
    if st.button("📊 Generar archivo CTG"):
        archivo_excel = exportar_excel(ficha_cb, fuente=fuente, tamaño=tamaño)
        nivel_tension = ficha_cb.get("Nivel de tensión (kV)", "XX")
        st.download_button(
            label="📥 Descargar archivo CTG en Excel",
            data=archivo_excel,
            file_name=f"CTG_{nivel_tension}kV.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
            

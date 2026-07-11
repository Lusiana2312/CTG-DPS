# empieza codigo
import streamlit as st
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime
import pandas as pd
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import textwrap
from PIL import Image as PILImage


################## CTG CIRCUIT BREAKER// INTERRUPTOR DE POTENCIA
def mostrar_app():

    st.title("📄 Generador de Ficha CTG")
    st.subheader("Interruptor de Potencia")

    # 1. Fabricante
    fabricante = "Indicar"
    st.text("### 🏢 Fabricante: " + fabricante)
    # 2. País
    pais = "Indicar"
    st.text("### País: " + pais)
    # 3. Referencia
    referencia = "Indicar"
    st.text("### Referencia: " + referencia)
    # 4. Norma de fabricación
    norma_fabricacion = "IEC 62271-100 / IEC 62271-110"
    st.markdown(f"**Norma de fabricación:** {norma_fabricacion}")
    # 5. Norma de calidad
    norma_calidad = "ISO 9001"
    st.markdown(f"**Norma de calidad:** {norma_calidad}")
    # 6. Medio de extinción
    medio_extincion = st.selectbox("Medio de extinción", ["Vacío", "SF6", "Aceite", "Aire comprimido"])
    # 7. Número de polos
    num_polos = 3
    st.text("Número de polos: {num_polos}")
    # 8. Número de cámaras por polo
    camaras_por_polo = "Indicar"
    st.text(f"### 🔢 Número de cámaras polo: " + camaras_por_polo)
    # 9. Tipo de ejecución
    tipo_ejecucion = st.selectbox("Tipo de ejecución", ["Exterior", "Interior"])
    # 10. Altura 
    altura_instalacion = st.number_input("Altura de instalación (m.s.n.m)", min_value=0, value=1000)

    # 11. Temperatura de operación
    st.markdown("### 🌡️ Temperatura de operación")
    temp_min = -5
    st.text(f"### Temperatura mínima anual (°C): {temp_min}")
    temp_max = +40
    st.text(f"### Temperatura máxima anual (°C): {temp_max}")
    temp_media = +35
    st.text(f"### Temperatura media (24 h) (°C): {temp_media}")
    
    # 12. Categoría de corrosión del ambiente
    categoria_corrosion = st.selectbox(
        "Categoría de corrosión del ambiente (ISO 12944-2 / ISO 9223)",
        options=["C1 - Muy baja", "C2 - Baja", "C3 - Media", "C4 - Alta", "C5 - Muy alta"]
    )
    # 13. Frecuencia
    frecuencia_asignada = "60 Hz"
    st.text(f"### Frecuencia asignada (fr): " + frecuencia_asignada)
    # 14. Tensión asignada Ur
    ur = st.selectbox("Tensión asignada (Ur)", options=["145 kV", "245 kV", "550 kV"])
    
    # 15. Tensión asignada a frecuencia industrial
    # Asignación automática de Ud según Ur
    ud_por_ur = {
        "145 kV": {"fase_tierra_ud": "750 kV", "entre_fases_ud": "750 kV", "interruptor_abierto_ud": "860 kV"},
        "245 kV": {"fase_tierra_ud": "460 kV", "entre_fases_ud": "460 kV", "interruptor_abierto_ud": "520 kV"},
        "550 kV": {"fase_tierra_ud": "830 kV", "entre_fases_ud": "830 kV", "interruptor_abierto_ud": "1150 kV"}
    }
    ud_valores = ud_por_ur.get(ur,{"fase_tierra_ud": "", "entre_fases_ud": "", "interruptor_abierto_ud": ""})
    st.markdown("#### Tensión asignada soportada a frecuencia industrial (Ud)")
    st.markdown(f"a) Fase-Tierra: **{ud_valores['fase_tierra_ud']}**")
    st.markdown(f"b) Entre fases: **{ud_valores['entre_fases_ud']}**")
    st.markdown(f"c) A través de interruptor abierto: **{ud_valores['interruptor_abierto_ud']}**")
    
    # 16. Tensión asignada a impulso maniobra
    # Asignación automática de Us por componente según Ur
    us_por_ur = {
        "145 kV": {"fase_tierra_us": "N.A.", "entre_fases_us": "N.A.", "interruptor_abierto_us": "N.A."},
        "245 kV": {"fase_tierra_us": "N.A.", "entre_fases_us": "N.A.", "interruptor_abierto_us": "N.A."},
        "550 kV": {"fase_tierra_us": "1300 kV", "entre_fases_us": "2210 kV", "interruptor_abierto_us": "1300 kV"}
    }
    us_valores = us_por_ur.get(ur, {"fase_tierra_us": "", "entre_fases_us": "", "interruptor_abierto_us": ""})
    st.markdown("#### Tensión asignada soportada a impulso de maniobra (Us)")
    st.markdown(f"a) Fase-Tierra: **{us_valores['fase_tierra_us']}**")
    st.markdown(f"b) Entre fases: **{us_valores['entre_fases_us']}**")
    st.markdown(f"c) A través de interruptor abierto: **{us_valores['interruptor_abierto_us']}**")

    
    # 17. Tensión asignada a impulso tipo rayo
    # Asignación automática de Up según Ur
    up_por_ur = {
        "145 kV": {"fase_tierra_up": "750 kV", "entre_fases_up": "750 kV", "interruptor_abierto_up": "860 kV"},
        "245 kV": {"fase_tierra_up": "1175 kV", "entre_fases_up": "1175 kV", "interruptor_abierto_up": "1175(+205)"},
        "550 kV": {"fase_tierra_up": "1800 kV", "entre_fases_up": "1800 kV", "interruptor_abierto_up": "1800 (+455) kV"}
    }
    up_valores = up_por_ur.get(ur, {"fase_tierra_up": "", "entre_fases_up": "", "interruptor_abierto_up": ""})
    st.markdown("#### Tensión asignada soportada a impulso tipo rayo (Up)")
    st.markdown(f"a) Fase-Tierra: **{up_valores['fase_tierra_up']}**")
    st.markdown(f"b) Entre fases: **{up_valores['entre_fases_up']}**")
    st.markdown(f"c) A través de interruptor abierto: **{up_valores['interruptor_abierto_up']}**")

    # 18. Corriente asignada
    ir_por_ur = {
        "145 kV": ["1200 A"],
        "245 kV": ["4000 A"],
        "550 kV": ["2500 A"]
    }
    # Mostrar opciones de Ir según Ur
    opciones_ir = ir_por_ur.get(ur, [])
    ir = st.selectbox("Corriente asignada en servicio continuo (Ir)", opciones_ir)
    
    # 19. Poder de corte asignado (Ics) según Ur
    ics_por_ur = {
        "145 kV": ["25 kA", "31.5 kA", "40 kA"],
        "245 kV": ["40 kA"],
        "550 kV": ["50 kA"]
    }
    # Mostrar opciones de Ics según Ur
    opciones_ics = ics_por_ur.get(ur, [])
    ics = st.selectbox("Poder de corte asignado en cortocircuito (Ics)", opciones_ics)

    # 20. Duración del cortocircuito asignado (Ics)
    duracion_ics = "1 s"
    st.text("### Duración del cortocircuito asignado (Ics): " + duracion_ics)
    
    # 21. Porcentaje de corriente aperiódica (%)
    porcentaje_ap = "Indicar"
    st.text("### Porcentaje de corriente aperiódica (%): " +  porcentaje_ap)

    # 22.Poder de cierre asignado en cortocircuito (Ip)
    st.markdown("**Poder de cierre asignado en cortocircuito (Ip):** 2.6 × Ics")

    # 23. Factor de primer polo según Ur
    factor_primer_polo = 1,5
    st.text(f"**Factor de primer polo:** {factor_primer_polo}")

    # 24. Tensión transitoria de restablecimiento asignada para fallas en bornes
    st.markdown("### ⚡ Tensión transitoria de restablecimiento asignada para fallas en bornes")
    # Mostrar los textos fijos
    st.markdown("**a)** Primera tensión de referencia (u1) kV: Indicar")
    st.markdown("**b)** Tiempo t1 ms: Indicar")
    st.markdown("**c)** Valor cresta del TTR (uc) kV: Indicar")
    st.markdown("**d)** Tiempo t2 ms: Indicar")
    st.markdown("**e)** Retardo td ms: Indicar")
    st.markdown("**f)** Tensión u’ kV: Indicar")
    st.markdown("**g)** Tiempo t’ ms: Indicar")
    st.markdown("**h)** Velocidad de crecimiento (u1 / t1) kV/ms: Indicar")
    
    # Asignar valores a las variables
    u1 = t1 = uc = t2 = td = u_prima = t_prima = vel_crecimiento = "Indicar"

    
    # 25. Características asignadas para fallas próximas en líneas
    st.markdown("### ⚡ Características asignadas para fallas próximas en líneas")

    # a) Características asignadas del circuito de alimentación
    st.markdown("#### a) Características asignadas del circuito de alimentación")
    st.markdown("**•)** Primera tensión de referencia (u1) kV: Indicar")
    st.markdown("**•)** Tiempo t1 ms: Indicar")
    st.markdown("**•)** Valor cresta del TTR (uc) kV: Indicar")
    st.markdown("**•)** Tiempo t2 ms: Indicar")
    st.markdown("**•)** Retardo td ms: Indicar")
    st.markdown("**•)** Tensión u’ kV: Indicar")
    st.markdown("**•)** Tiempo t’ ms: Indicar")
    st.markdown("**•)** Velocidad de crecimiento (u1 / t1) kV/ms: Indicar")
    u1_alimentacion = t1_alimentacion = uc_alimentacion = t2_alimentacion = td_alimentacion = u_prima_alimentacion = t_prima_alimentacion = vel_crecimiento_alimentacion = "Indicar"

    # b) Características asignadas de la línea
    st.markdown("#### b) Características asignadas de la línea")
    st.markdown("• Impedancia de onda asignada (Z)")
    st.markdown("• Factor de cresta asignada (k)")
    st.markdown("• Factor de TCTR (s)")
    st.markdown("• Retardo (tdl)")
    z_linea = k_linea = s_linea = tdl_linea = "Indicar"


    
    # 26. Característica de TRV de pequeñas corrientes inductivas según IEC 62271-110
    st.markdown("### ⚡ Característica de TRV de pequeñas corrientes inductivas según IEC 62271-110")
    
    # Rangos de referencia según Ur
    rangos_trv = {
        "145 kV": {"Uc": "225 kV", "t3_1": "105 µs", "t3_2": "187 µs"},
        "245 kV": {"Uc": "380 kV", "t3_1": "167 µs", "t3_2": "297 µs"},
        "550 kV": {"Uc": "1240 kV", "t3_1": "300 µs", "t3_2": "536 µs"}
    }
    
    # Obtener valores según Ur
    valores_trv = rangos_trv.get(ur, {"Uc": "Indicar", "t3_1": "Indicar", "t3_2": "Indicar"})
    
    # Mostrar los valores como texto
    st.markdown(f"**a)** Valor mínimo pico de TRV Uc: {valores_trv['Uc']}")
    st.markdown(f"**b)** Tiempo máximo t₃ Load circuit 1: {valores_trv['t3_1']}")
    st.markdown(f"**c)** Tiempo máximo t₃ Load circuit 2: {valores_trv['t3_2']}")
    
    # Asignar variables para exportación o uso posterior
    uc_trv = valores_trv["Uc"]
    t3_1 = valores_trv["t3_1"]
    t3_2 = valores_trv["t3_2"]
    

    # 27. Tiempo de arco mínimo ante pequeñas corrientes inductivas
    st.markdown("### ⏱️ Tiempo de arco mínimo ante pequeñas corrientes inductivas")
    tiempo_arco_minimo = "<=5 ms"
    st.text("### Tiempo de arco mínimo (Minimum Arcing Time): " + tiempo_arco_minimo)

    ################################################

    # 28. Número de corte λ ("Chopping Number λ")
    st.markdown("### 🔢 Número de corte λ (Chopping Number λ)")
    if ur == "145 kV":
        numero_corte_lambda = "<=3x10^4"
    elif ur == "245 kV":
        numero_corte_lambda = "<=5x10^4"
    elif ur == "550 kV":
        numero_corte_lambda = "<=7x10^4"
    else:
        numero_corte_lambda = "Indicar"
    
    # Mostrar el resultado como texto
    st.markdown(f"**Número de corte λ:** {numero_corte_lambda}")

    
    # 29. Secuencia de maniobras asignada
    st.markdown("### 🔁 Secuencia de maniobras asignada")
    secuencia_maniobras = "O-0,3s-CO-3min-CO"
    st.text("Secuencia de maniobras asignada: " + secuencia_maniobras)

    # 30. Poder de corte en discordancia de fases (Id)
    st.markdown("### ⚡ Poder de corte en discordancia de fases (Id)")
    
    st.markdown("**a)** Primera tensión de referencia (u1) [Id]: Indicar")
    st.markdown("**b)** Tiempo t1 [Id]: Indicar")
    st.markdown("**c)** Valor cresta del TTR (uc) [Id]: Indicar")
    st.markdown("**d)** Tiempo t2 [Id]: Indicar")
    st.markdown("**e)** Velocidad de crecimiento (u1 / t1) [Id]: Indicar")
    
    # Asignación fija de valores
    id_u1 = id_t1 = id_uc = id_t2 = id_vel_crecimiento = "Indicar"
    
    # 31. Apertura de líneas en vacío
    st.markdown("### ⚡ Apertura de líneas en vacío")
    st.markdown("**a)** Poder de corte asignado (Ir) [Apertura de líneas en vacío]: Indicar")
    st.markdown("**b)** Sobretensión de maniobra presente: Indicar")
    # Asignación fija de valores
    ir_apertura_linea = sobretension_maniobra = "Indicar"

    # 32. Apertura de corrientes inductivas pequeñas
    st.markdown("### ⚡ Apertura de corrientes inductivas pequeñas")
    
    # Campo editable (se mantiene igual)
    apertura_inductiva = st.selectbox("¿Aplica apertura de corrientes inductivas pequeñas?", ["Sí", "No"])
    # Campos fijos con valor "Indicar"
    st.markdown("**a)** Poder de corte asignado [corrientes inductivas pequeñas]: Indicar")
    st.markdown("**b)** Sobretensión de maniobra máxima: Indicar")
    # Asignación fija de valores
    ir_inductiva = sobretension_inductiva = "Indicar"
    
    # 33. Número de operaciones mecánicas
    st.markdown("### ⚙️ Número de operaciones mecánicas")
    num_operaciones_mecanicas = st.selectbox("Número de operaciones mecánicas", ["M1", "M2", "M3"])

    # 34. Probabilidad de reencendido
    st.markdown("### 🔄 Probabilidad de reencendido")
    probabilidad_reencendido = st.selectbox("Probabilidad de reencendido", ["C1", "C2"])

    # 35. Máxima diferencia de tiempo entre contactos de diferente polo
    st.markdown("### ⏱️ Máxima diferencia de tiempo entre contactos de diferente polo")
    st.markdown("**Máxima diferencia de tiempo entre contactos de diferente polo al tocarse durante un cierre o al separarse durante una apertura:** Indicar")
    # Asignación fija
    diferencia_tiempo_contactos = "Indicar"
    
    # 36. Maniobra de apertura
    st.markdown("### 🔧 Maniobra de apertura")
    st.markdown("**a)** Tiempo de apertura: Indicar")
    st.markdown("**b)** Tiempo de arco: Indicar")
    st.markdown("**c)** Tiempo máximo de corte asignado: 40 ms")
    tiempo_apertura = "Indicar"
    tiempo_arco = "Indicar"
    tiempo_max_corte = "40 ms"
    
    # 37. Tiempo muerto
    st.markdown("### ⏳ Tiempo muerto")
    st.markdown("**Tiempo muerto:** Indicar")
    tiempo_muerto = "Indicar"
    
    # 38. Maniobra de cierre
    st.markdown("### 🔧 Maniobra de cierre")
    
    st.markdown("**a)** Tiempo de establecimiento: Indicar")
    st.markdown("**b)** Tiempo de prearco: Indicar")
    st.markdown("**c)** Tiempo de cierre: Indicar")
    tiempo_establecimiento = tiempo_prearco = tiempo_cierre = "Indicar"

    # 39. Gas SF6 - Interruptor
    st.markdown("### 🧪 Gas SF₆ – Interruptor")
    st.markdown("**a)** Presión de gas asignada para maniobra (Pob): Indicar")
    st.markdown("**b)** Presión de gas asignada para el corte (Pcb): Indicar")
    presion_maniobra = presion_corte = "Indicar"

    # 40. Volumen total de SF6 por polo a 0,1 MPa
    st.markdown("### 🧪 Volumen total de SF₆ por polo a 0,1 MPa")
    st.markdown("**Volumen total de SF₆ por polo a 0,1 MPa:** Indicar")
    volumen_sf6 = "Indicar"

    # 41. Pérdida máxima de SF6 por año (valor fijo)
    st.markdown("### Pérdida máxima de SF₆ por año")
    perdida_sf6 = "≤ 0.5%"
    st.markdown(f"**Pérdida máxima de SF₆ por año:** {perdida_sf6}")
    
    # 42. Resistencia máxima entre terminales
    st.markdown("### Resistencia máxima entre terminales")
    st.markdown("**Resistencia máxima entre terminales (μΩ):** Indicar")
    resistencia_max_terminales = "Indicar"

    ################### PREGUNTAR
    #  43. Capacitancia
    st.markdown("### Capacitancia")
    cap_entre_contactos_con_resistencia = st.markdown("a) Entre contactos abiertos - Con resistencia de preinserción (pF): Indicar")
    cap_entre_contactos_sin_resistencia = st.markdown("a) Entre contactos abiertos - Sin resistencia de preinserción (pF): Indicar")
    cap_entre_contactos_tierra = st.markdown("b) Entre contactos y tierra (pF): Indicar")
    cap_condensador_gradiente = st.markdown("c) Condensador de gradiente (***) (pF): Indicar")
    cap_entre_contactos_con_resistencia = cap_entre_contactos_sin_resistencia = cap_entre_contactos_tierra = cap_condensador_gradiente = cap_entre_contactos_con_resistencia = "Indicar"

    # 44. Material de los empaques
    st.markdown("### Material de los empaques")
    st.markdown("**Material de los empaques:** Indicar")
    material_empaques = "Indicar"

    # 45. Operación con mando sincronizado
    st.markdown("### Operación con mando sincronizado")
    mando_sincronizado = st.radio("¿Operación con mando sincronizado?", ["Sí", "No"])

    # 46. Resistencia de preinserción
    st.markdown("### Resistencia de preinserción")
    resistencia_preinsercion = st.radio("¿Resistencia de preinserción?", ["Sí", "No"])

    # 47. Distancia mínima en aire
    st.markdown("### Distancia mínima en aire")
    st.markdown("**a)** Entre polos (mm): Indicar")
    st.markdown("**b)** A tierra (mm): Indicar")
    st.markdown("**c)** A través del polo (mm): Indicar")
    distancia_entre_polos = distancia_a_tierra = distancia_a_traves_polo = "Indicar"

    # 48. Clase de severidad de contaminación del sitio (SPS)
    st.markdown("### Clase de severidad de contaminación del sitio (SPS)")
    sps_clase = st.selectbox(
        "Clase de severidad de contaminación del sitio (SPS)",
        ["Ligera", "Media", "Pesada", "Muy pesada"]
    )

    # 49. Distancia mínima de fuga
    st.markdown("### Distancia mínima de fuga")
    distancia_minima_fuga = st.text_input("Distancia mínima de fuga (mm)")

    # 50. Datos sísmicos
    st.markdown("### Datos sísmicos")
    st.markdown("**Desempeño sísmico según IEEE-693-Vigente:** Alto (0,5g)")
    st.markdown("**a)** Frecuencia natural de vibración (Hz): Indicar")
    st.markdown("**b)** Coeficiente de amortiguamiento crítico (%): Indicar")
    desempeno_sismico_ieee = "Alto (0,5g)"
    frecuencia_natural_vibracion = coef_amortiguamiento_critico = "Indicar"

    # 51. Cargas admisibles en bornes
    st.markdown("### Cargas admisibles en bornes")
    if ur == "145 kV":
        carga_estatica_admisible = "1000 N"
        carga_dinamica_admisible = "2000 N"
    elif ur == "245 kV":
        carga_estatica_admisible = "1500 N"
        carga_dinamica_admisible = "4000 N"
    elif ur == "550 kV":
        carga_estatica_admisible = "2000 N"
        carga_dinamica_admisible = "5500 N"
    else:
        carga_estatica_admisible = carga_dinamica_admisible = "Indicar"
    st.markdown(f"**a)** Carga estática admisible (N): {carga_estatica_admisible}")
    st.markdown(f"**b)** Carga dinámica admisible (N): {carga_dinamica_admisible}")

    # 52. Fuerzas asociadas a la operación del equipo
    st.markdown("### Fuerzas asociadas a la operación del equipo")
    st.markdown("**a)** Fuerza vertical (N): Indicar")
    st.markdown("**b)** Fuerza horizontal (N): Indicar")
    fuerza_vertical = fuerza_horizontal = "Indicar"

    # 53. Masa neta de un polo completo con estructura
    st.markdown("### Masa neta de un polo completo con estructura")
    st.markdown("** Masa neta de un polo completo con estructura (kg): Indicar")
    masa_neta_polo = "Indicar"

    # 54. Dimensiones para transporte
    st.markdown("### Dimensiones para transporte")
    st.markdown("Dimensiones para transporte (Alto x Ancho x Largo) [mm]: Indicar")
    dimensiones_transporte = "Indicar"
    
    # 55. Datos adicionales para transporte
    st.markdown("Masa neta para transporte (kg): Indicar")
    masa_neta_transporte = "Indicar"
    
    # 56. Volumen
    st.markdown("Volumen total para transporte (m³): Indicar")
    volumen_total_transporte = "Indicar"
    
    # 57. Campo eléctrico
    campo_electrico_1m = "<= 8,33"
    st.markdown("Campo eléctrico a 1 metro de separación del piso (kV/m)" + campo_electrico_1m)
    
    # BOTÓN PARA GENERAR FICHA
    ficha_cb = {
        "Fabricante": fabricante,
        "País": pais,
        "Referencia": referencia,
        "Norma de fabricación": norma_fabricacion,
        "Norma de calidad": norma_calidad,
        "Medio de extinción": medio_extincion,
        "Número de polos": num_polos,
        "Número de cámaras por polo": camaras_por_polo,
        "Tipo de ejecución": tipo_ejecucion,
        "Altura de instalación (m.s.n.m)": altura_instalacion,
        "Temperatura mínima anual (°C)": temp_min,
        "Temperatura máxima anual (°C)": temp_max,
        "Temperatura media (24 h) (°C)": temp_media,
        "Fecha de registro": datetime.now().strftime("%Y-%m-%d"),
        "Categoría de corrosión del ambiente": categoria_corrosion,
        "Frecuencia asignada (fr)": frecuencia_asignada,
        "Tensión asignada (Ur) [kV]": ur,
        "Ud - Fase-Tierra [kV]": ud_valores["fase_tierra_ud"],
        "Ud - Entre fases [kV]": ud_valores["entre_fases_ud"],
        "Ud - A través de interruptor abierto [kV]": ud_valores["interruptor_abierto_ud"],
        "Us - Fase-Tierra [kV]": us_valores["fase_tierra_us"],
        "Us - Entre fases [kV]": us_valores["entre_fases_us"],
        "Us - A través de interruptor abierto [kV]": us_valores["interruptor_abierto_us"],
        "Up - Fase-Tierra [kV]": up_valores["fase_tierra_up"],
        "Up - Entre fases [kV]": up_valores["entre_fases_up"],
        "Up - A través de interruptor abierto [kV]": up_valores["interruptor_abierto_up"],
        "Corriente asignada en servicio continuo (Ir)": ir,
        "Poder de corte asignado en cortocircuito (Ics)": ics,
        "Duración del cortocircuito asignado (Ics)": duracion_ics,
        "Porcentaje de corriente aperiódica (%)": porcentaje_ap,
        "Poder de cierre asignado en cortocircuito (Ip)": "2.6 × Ics",
        "Factor de primer polo": factor_primer_polo,
        "TTR - Primera tensión de referencia (u1)": u1,
        "TTR - Tiempo t1": t1,
        "TTR - Valor cresta del TTR (uc)": uc,
        "TTR - Tiempo t2": t2,
        "TTR - Retardo td": td,
        "TTR - Tensión u’": u_prima,
        "TTR - Tiempo t’": t_prima,
        "TTR - Velocidad de crecimiento (u1 / t1)": vel_crecimiento,
        "Fallas próximas - u1 alimentación": u1_alimentacion,
        "Fallas próximas - t1 alimentación": t1_alimentacion,
        "Fallas próximas - uc alimentación": uc_alimentacion,
        "Fallas próximas - t2 alimentación": t2_alimentacion,
        "Fallas próximas - td alimentación": td_alimentacion,
        "Fallas próximas - u’ alimentación": u_prima_alimentacion,
        "Fallas próximas - t’ alimentación": t_prima_alimentacion,
        "Fallas próximas - velocidad crecimiento alimentación": vel_crecimiento_alimentacion,
        "Fallas próximas - impedancia de onda (Z)": z_linea,
        "Fallas próximas - factor de cresta (k)": k_linea,
        "Fallas próximas - factor de TCTR (s)": s_linea,
        "Fallas próximas - retardo (tdl)": tdl_linea,
        "TRV - Valor mínimo pico de TRV Uc": uc_trv,
        "TRV - Tiempo máximo t₃ Load circuit 1": t3_1,
        "TRV - Tiempo máximo t₃ Load circuit 2": t3_2,
        "Tiempo de arco mínimo (Minimum Arcing Time)": tiempo_arco_minimo,
        "Número de corte λ (Chopping Number λ)": numero_corte_lambda,
        "Secuencia de maniobras asignada": secuencia_maniobras,
        "Poder de corte en discordancia de fases - u1": id_u1,
        "Poder de corte en discordancia de fases - t1": id_t1,
        "Poder de corte en discordancia de fases - uc": id_uc,
        "Poder de corte en discordancia de fases - t2": id_t2,
        "Poder de corte en discordancia de fases - velocidad de crecimiento (u1 / t1)": id_vel_crecimiento,
        "Apertura de líneas en vacío - Poder de corte asignado (Ir)": ir_apertura_linea,
        "Apertura de líneas en vacío - Sobretensión de maniobra presente": sobretension_maniobra,
        "Apertura de corrientes inductivas pequeñas": apertura_inductiva,
        "Apertura inductiva - Poder de corte asignado": ir_inductiva,
        "Apertura inductiva - Sobretensión de maniobra máxima": sobretension_inductiva,
        "Número de operaciones mecánicas": num_operaciones_mecanicas,
        "Probabilidad de reencendido": probabilidad_reencendido,
        "Máxima diferencia de tiempo entre contactos de diferente polo": diferencia_tiempo_contactos,
        "Maniobra de apertura - Tiempo de apertura": tiempo_apertura,
        "Maniobra de apertura - Tiempo de arco": tiempo_arco,
        "Maniobra de apertura - Tiempo máximo de corte asignado": tiempo_max_corte,
        "Tiempo muerto": tiempo_muerto,
        "Maniobra de cierre - Tiempo de establecimiento": tiempo_establecimiento,
        "Maniobra de cierre - Tiempo de prearco": tiempo_prearco,
        "Maniobra de cierre - Tiempo de cierre": tiempo_cierre,
        "Gas SF6 - Presión de maniobra (Pob)": presion_maniobra,
        "Gas SF6 - Presión de corte (Pcb)": presion_corte,
        "Volumen total de SF6 por polo a 0,1 MPa": volumen_sf6,
        "Pérdida máxima de SF6 por año": perdida_sf6,
        "Resistencia máxima entre terminales (μΩ)": resistencia_max_terminales,
        "Capacitancia - Entre contactos abiertos con resistencia de preinserción (pF)": cap_entre_contactos_con_resistencia,
        "Capacitancia - Entre contactos abiertos sin resistencia de preinserción (pF)": cap_entre_contactos_sin_resistencia,
        "Capacitancia - Entre contactos y tierra (pF)": cap_entre_contactos_tierra,
        "Capacitancia - Condensador de gradiente (***) (pF)": cap_condensador_gradiente,
        "Material de los empaques": material_empaques,
        "Operación con mando sincronizado": mando_sincronizado,
        "Resistencia de preinserción": resistencia_preinsercion,
        "Distancia mínima en aire - Entre polos (mm)": distancia_entre_polos,
        "Distancia mínima en aire - A tierra (mm)": distancia_a_tierra,
        "Distancia mínima en aire - A través del polo (mm)": distancia_a_traves_polo,
        "Clase de severidad de contaminación del sitio (SPS)": sps_clase,
        "Distancia mínima de fuga (mm)": distancia_minima_fuga,
        "Desempeño sísmico según IEEE-693-Vigente (**)": desempeno_sismico_ieee,
        "Frecuencia natural de vibración (Hz)": frecuencia_natural_vibracion,
        "Coeficiente de amortiguamiento crítico (%)": coef_amortiguamiento_critico,
        "Cargas admisibles en bornes - Carga estática admisible (N)": carga_estatica_admisible,
        "Cargas admisibles en bornes - Carga dinámica admisible (N)": carga_dinamica_admisible,
        "Fuerzas asociadas a la operación del equipo - Vertical (N)": fuerza_vertical,
        "Fuerzas asociadas a la operación del equipo - Horizontal (N)": fuerza_horizontal,
        "Masa neta de un polo completo con estructura (kg)": masa_neta_polo,
        "Dimensiones para transporte (Alto x Ancho x Largo) [mm]": dimensiones_transporte,
        "Masa neta para transporte (kg)": masa_neta_transporte,
        "Volumen total para transporte (m³)": volumen_total_transporte,
        "Campo eléctrico a 1 metro de separación del piso (kV/m)": campo_electrico_1m
            
    }
    
    # 📤 Función para exportar Excel con estilo personalizado
    def exportar_excel(datos, fuente="Calibri", tamaño=9):
        # Diccionario de unidades (puedes ampliarlo según tus campos)
        unidades = {
            "Tensión asignada (Ur) [kV]": "kV",
            "Altura de instalación (m.s.n.m)": "m.s.n.m",
            "Temperatura mínima anual (°C)": "°C",
            "Temperatura máxima anual (°C)": "°C",
            "Temperatura media (24 h) (°C)": "°C",
            "Frecuencia asignada (fr)": "Hz",
            "Corriente asignada en servicio continuo (Ir)": "A",
            "Poder de corte asignado en cortocircuito (Ics)": "kA",
            "Duración del cortocircuito asignado (Ics)": "s",
            "Porcentaje de corriente aperiódica (%)": "%",
            "Distancia mínima en aire - Entre polos (mm)": "mm",
            "Distancia mínima de fuga (mm)": "mm",
            "Campo eléctrico a 1 metro de separación del piso (kV/m)": "kV/m",
            "Masa neta para transporte (kg)": "kg",
            "Volumen total para transporte (m³)": "m³",
            "Dimensiones para transporte (Alto x Ancho x Largo) [mm]": "mm",
            "Masa neta de un polo completo con estructura (kg)": "kg"
            # Añade más unidades según tus campos
        }
    
        # Crear DataFrame con estructura personalizada
        df = pd.DataFrame([
            {
                "ÍTEM": i + 1,
                "DESCRIPCIÓN": campo,
                "UNIDAD": unidades.get(campo, ""),
                "REQUERIDO": valor,
                "OFRECIDO": ""  # Columna vacía para completar manualmente
            }
            for i, (campo, valor) in enumerate(datos.items())
        ])
    
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="CTG", startrow=6)
            wb = writer.book
            ws = writer.sheets["CTG"]
            ws.print_title_rows = '1:7'
            ws.print_area = f"A1:E{ws.max_row}"

            
            # 🖼️ Insertar imagen del logo (opcional)
            logo_path = "siemens_logo.png"
            try:
                # Abrimos con PIL para procesar cualquier formato (incluyendo webp disfrazado)
                img_aux = PILImage.open(logo_path)
                img_converted = BytesIO()
                img_aux.save(img_converted, format="PNG") # Forzamos PNG
                img_converted.seek(0)
                
                # Insertamos en Excel
                img = Image(img_converted)
                img.width = 280
                img.height = 90
                ws.add_image(img, "C1")
            except Exception as e:
                st.warning(f"⚠️ No se pudo insertar el logo: {e}. El archivo se generará sin imagen.")
    
            # 🟪 Caja de título
            ws.merge_cells("A2:E4")
            cell = ws.cell(row=2, column=1)
            cell.value = "FICHA TÉCNICA INTERRUPTOR DE POTENCIA"
            cell.font = Font(name=fuente, bold=True, size=14, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
            # 🏷️ Subtítulo técnico
            ws.merge_cells("A5:D5")
            ws["A5"] = f"CARACTERÍSTICAS GARANTIZADAS"
            ws["A5"].font = Font(name=fuente, bold=True, size=12)
            ws["A5"].alignment = Alignment(horizontal="center")
    
            # 🎨 Encabezados con estilo
            header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            header_font = Font(name=fuente, size=tamaño, color="FFFFFF", bold=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
            for col_num in range(1, 6):
                cell = ws.cell(row=6, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
    
            # 📐 Ajuste de columnas
            ws.column_dimensions["A"].width = 4
            ws.column_dimensions["B"].width = 50
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 12
    
            
            
            # 📋 Formato de filas con fuente personalizada y ajuste dinámico de altura
            for row in ws.iter_rows(min_row=7, max_row=ws.max_row, max_col=5):
                max_lines = 1  # Mínimo una línea por celda
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.font = Font(name=fuente, size=tamaño)
            
                    # Estimar número de líneas necesarias si el contenido es texto
                    if cell.value and isinstance(cell.value, str):
                        # Ajusta el ancho según la columna (por ejemplo, columna B tiene 55 caracteres de ancho)
                        if cell.column_letter == "B":
                            wrapped = textwrap.wrap(cell.value, width=55)
                            max_lines = max(max_lines, len(wrapped))
            
                # Ajustar altura de la fila según el contenido más largo
                ws.row_dimensions[row[0].row].height = max_lines * 15  # 15 puntos por línea aprox.
            
                # Alineación horizontal para columnas específicas
                row[0].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                row[2].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                row[3].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                row[4].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
                
        output.seek(0)
        return output
    
    # 📥 Botón para generar y descargar
    fuente = "Calibri"
    tamaño = 9
    if st.button("📊 Generar archivo CTG"):
        archivo_excel = exportar_excel(ficha_cb, fuente=fuente, tamaño=tamaño)
        nivel_tension = ficha_cb.get("Nivel de tensión (kV)", "XX")
        st.download_button(
            label="📥 Descargar archivo CTG en Excel",
            data=archivo_excel,
            file_name=f"CTG_{nivel_tension}kV.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
            
